"""Generate the SSOT Spec CSV + multi-tab XLSX from a single declarative source.

Run: python backend/scripts/gen_ssot_spec_csv.py
Output:
  docs/ssot-spec.csv   — single-tab CSV (header + description + consumer + 5 SKUs)
  docs/ssot-spec.xlsx  — 2-tab Excel: "Read Me" + "SKU Operational Database"
                         (upload to Drive; Google converts to multi-tab Sheet)

Wide format. Row 1 = column headers. Row 2 = source/how-to-fill descriptions.
Row 3 = consumers (who reads this column).
Rows 4-8 = 5 example SKUs filled with real OCR data from catalogue_items.

Cell value conventions:
  <value>       → known value (from OCR, master list, or computed)
  incomplete    → required but not yet collected (OCR didn't extract, pipeline not built)
  N/A (rule)    → not applicable for this SKU per a known business rule
"""
import csv, os
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

# ─── Consumer labels ───────────────────────────────────────────────────────
# Coloured square emoji matches the consumer chip colours on /ssot-spec.
# AM and PS both feed the same Approval Matrix conceptually, but are kept
# visually distinct as two workstreams because their data needs may differ.
CONSUMER_LABELS = {
    "AM":           "🟧 AM (Approval Matrix)",
    "PS":           "🟦 PS (Product Selection)",
    "IMS":          "🟪 IMS (Rosetta IMS)",
    "SHOPIFY_OUT":  "🟩 → Shopify (push)",
    "DAYSMART_OUT": "🟨 → DaySmart (push)",
}

def consumers_text(keys):
    """Multi-line cell content: each consumer on its own line."""
    return "\n".join(CONSUMER_LABELS[k] for k in keys)

# ─── Column metadata: (key, description, consumers_list) in order ─────────
# If consumers_list is empty, the column has no downstream use — REMOVE it.
COLUMNS_RAW = [
    ("sku_id",                          "⚪ AUTO: 8-digit code from RECORD | SKU MASTER LIST"),
    ("sku_name",                        "🟡 OCR+HUMAN: catalogue → /data-review (OCR fill ~99%)"),
    ("brand",                           "🟡 OCR+HUMAN: catalogue → /data-review (OCR fill 35% — gap)"),
    ("category",                        "🟡 OCR+HUMAN: catalogue: Medicine / Food / Supplement / Preventative / Pet Hygiene / Not-For-Sale"),
    ("sub_category",                    "🟡 OCR+HUMAN: catalogue: Vaccine/Chemo/Antibiotics/Wet/Dry/Joint/etc."),
    ("species",                         "🟡 OCR+HUMAN: catalogue: Dog / Cat / Both / Other"),
    ("storage_rule",                    "🟢 MANUAL: clinic_only (Medicine usually) / any"),
    ("status",                          "🟢 MANUAL: ACTIVE / INACTIVE / DISCONTINUED"),
    ("hero_sku",                        "🟢 MANUAL: Yes if QGB/SGB flagged for WOC priority"),
    ("clinic_needs_it",                 "🟢 MANUAL (NEW per 29 May): Yes if Dr James says clinic absorbs leftover units"),
    ("listed_on_hktv",                  "🔵 TECH: HKTV merchant export"),
    ("listed_on_shopify",               "🔵 TECH: Shopify Admin API"),
    ("primary_supplier_id",             "🟡 OCR+HUMAN: catalogue: FK to Suppliers table"),
    ("alternative_supplier_ids",        "🟡 OCR+HUMAN: catalogue: comma-separated supplier FKs"),
    ("supplier_sku_code",               "🟡 OCR+HUMAN: catalogue → /data-review (OCR fill 77%)"),
    ("supplier_barcode",                "🟡 OCR+HUMAN: catalogue (OCR fill 0% — gap, catalogues don't print)"),
    ("supplier_moq_cached",             "⚪ AUTO: copied from Suppliers.moq_hkd"),
    ("catalogue_cost",                  "🟡 OCR+HUMAN · NUMBER · HKD per supplier pack as printed (OCR fill 87%). SOURCE for basic_unit_cost formula."),
    ("basic_unit_cost",                 "🟡 FORMULA · =catalogue_cost / units_per_pack. Auto-derived per MOU. Never hardcode."),
    ("daysmart_last_invoice_cost",      "🔵 TECH · NUMBER · HKD per MOU from DaySmart invoice (PIPELINE NOT BUILT YET)"),
    ("shopify_last_invoice_cost",       "🔵 TECH · NUMBER · HKD per MOU from Shopify invoice/PO (PIPELINE NOT BUILT YET)"),
    ("mbb_pack_cost",                   "🟡 OCR+HUMAN · NUMBER · HKD per supplier pack at MBB tier (raw, before per-unit divide). SOURCE for mbb_cost_per_unit formula."),
    ("mbb_cost_per_unit",               "🟡 FORMULA · =mbb_pack_cost / units_per_pack. Auto-derived per MOU at MBB. Never hardcode."),
    ("mbb_min_qty",                     "🟡 OCR+HUMAN · INTEGER · supplier packs to hit MBB tier (OCR fill 18%)"),
    ("mbb_tier_structure",              "🟡 OCR+HUMAN · JSON · [{qty,price}] array of all bulk tiers. Today often free text — gap."),
    ("cost_source",                     "⚪ AUTO · ENUM · catalogue (OCR, top) > invoice_matched > po_issued > manual > sheet (seed)"),
    ("cost_last_reviewed_at",           "⚪ AUTO · ISO DATE · stamped on /data-review approval"),
    ("minimum_operating_unit",          "🟡 OCR+HUMAN: catalogue: smallest customer-buyable unit"),
    ("uom_sell_unit",                   "🔵 TECH · ENUM · sell unit name (tablet / mL / bottle / tube / pot)"),
    ("minimum_purchase_qty",            "🟡 OCR+HUMAN · INTEGER · MOU per minimum supplier pack (e.g. 150 tabs/bottle → 150). NEW per 29 May."),
    ("units_per_pack",                  "🟡 OCR+HUMAN · INTEGER · MOU per supplier case (OCR fill 91%, semantics confused for liquids)"),
    ("weight_grams",                    "🟡 OCR+HUMAN · NUMBER · grams per MOU (drives SF Express fee). NOT EXTRACTED YET — gap."),
    ("risk_acceptance",                 "🟢 MANUAL (NEW per 29 May): pass_to_customer / we_take_risk / clinic_absorbs"),
    ("shopify_minimum_qty",             "🟢 MANUAL: derived from MPQ + risk → Shopify min setting"),
    ("shopify_multiples",               "🟢 MANUAL: derived → Shopify multiples setting"),
    ("sell_price_clinic",               "🔵 TECH: DaySmart Vet POS price list"),
    ("sell_price_shopify",              "🔵 TECH: Shopify Admin API"),
    ("sell_price_hktv",                 "🔵 TECH: HKTV merchant export"),
    ("dispensing_fee_clinic",           "🔵 TECH: DaySmart pull. Human judgment → entered in DaySmart → pulled into SSOT. HKD ADDED ON TOP of clinic sell price. 0 = doesn't apply. Lifts true clinic GP."),
    ("dispensing_fee_ecommerce",        "🔵 TECH: Shopify pull (or IMS config). Human judgment → entered in Shopify → pulled into SSOT. HKD ADDED ON TOP of online sell price. Usually differs from clinic fee."),
    ("channel_fee_pct_hktv",            "🔵 TECH: HKTV fee schedule"),
    ("stock_clinic",                    "🔵 TECH: DaySmart inventory report"),
    ("stock_warehouse",                 "🔵 TECH: ShopToPlus app via Shopify"),
    ("demand_120d_daysmart",            "🔵 TECH: DaySmart 120-day sales report"),
    ("demand_120d_shopify",             "🔵 TECH: Shopify Admin API"),
    ("demand_120d_hktv",                "🔵 TECH: HKTV merchant export"),
    ("unfulfilled_jit",                 "🔵 TECH: Shopify paid+unfulfilled count"),
    ("upcoming_14d_autoship",           "🔵 TECH: Shopify subscription module"),
    ("competitor_lowest_price",         "🔵 TECH: HKTV scrape (or manual until automated)"),
    ("competitor_source_url",           "🟢 MANUAL: URL of competitor listing"),
    ("competitor_last_reviewed",        "🟢 MANUAL: date competitor price last verified"),
    ("rrp",                             "🟡 OCR+HUMAN: catalogue RRP (NOT EXTRACTED YET — gap)"),
    ("sf_express_fee",                  "🔵 TECH: weight_grams × SF Express rate table"),
    ("expiration_date",                 "🟢 MANUAL: from supplier Whatsapp / batch docs"),
    ("bulk_purchase_benefit_available", "🟢 MANUAL: Yes/No — supplier currently offering bulk rebate"),
    ("last_updated_by",                 "⚪ AUTO: stamped on write"),
    ("last_updated_at",                 "⚪ AUTO: stamped on write"),
]

# ─── Consumers per column ──────────────────────────────────────────────────
# Who downstream actually reads this column? If empty list → column has no
# consumer → it will be REMOVED from the output sheet (warned at gen time).
CONSUMERS_BY_KEY = {
    "sku_id":                          ["IMS", "AM", "PS", "SHOPIFY_OUT", "DAYSMART_OUT"],
    "sku_name":                        ["IMS", "AM", "PS"],
    "brand":                           ["IMS", "AM", "PS"],
    "category":                        ["AM", "PS", "IMS"],
    "sub_category":                    ["AM", "PS", "IMS"],
    "species":                         ["PS", "IMS"],
    "storage_rule":                    ["IMS"],
    "status":                          ["AM", "PS", "IMS"],
    "hero_sku":                        ["AM", "PS", "IMS"],
    "clinic_needs_it":                 ["AM", "PS"],
    "listed_on_hktv":                  ["PS", "IMS"],
    "listed_on_shopify":               ["PS", "IMS", "SHOPIFY_OUT"],
    "primary_supplier_id":             ["AM", "PS", "IMS"],
    "alternative_supplier_ids":        ["PS", "IMS"],
    "supplier_sku_code":               ["AM", "IMS"],
    "supplier_barcode":                ["IMS"],
    "supplier_moq_cached":             ["AM", "PS"],
    "basic_unit_cost":                 ["AM", "PS", "IMS"],
    "catalogue_cost":                  ["IMS"],
    "daysmart_last_invoice_cost":      ["AM", "IMS"],
    "shopify_last_invoice_cost":       ["AM", "IMS"],
    "mbb_pack_cost":                   ["AM", "PS", "IMS"],
    "mbb_cost_per_unit":               ["AM", "PS", "IMS"],
    "mbb_min_qty":                     ["AM", "PS", "IMS"],
    "mbb_tier_structure":              ["PS", "IMS"],
    "cost_source":                     ["IMS"],
    "cost_last_reviewed_at":           ["AM", "IMS"],
    "minimum_operating_unit":          ["PS", "SHOPIFY_OUT"],
    "uom_sell_unit":                   ["IMS", "SHOPIFY_OUT"],
    "minimum_purchase_qty":            ["PS", "AM", "SHOPIFY_OUT"],
    "units_per_pack":                  ["AM", "PS", "IMS"],
    "weight_grams":                    ["AM", "PS", "IMS"],
    "risk_acceptance":                 ["PS", "SHOPIFY_OUT"],
    "shopify_minimum_qty":             ["SHOPIFY_OUT"],
    "shopify_multiples":               ["SHOPIFY_OUT"],
    "sell_price_clinic":               ["AM", "PS", "IMS"],
    "sell_price_shopify":              ["AM", "PS", "IMS"],
    "sell_price_hktv":                 ["AM", "PS", "IMS"],
    "dispensing_fee_clinic":           ["AM", "PS", "IMS"],
    "dispensing_fee_ecommerce":        ["AM", "PS", "IMS"],
    "channel_fee_pct_hktv":            ["AM", "PS"],
    "stock_clinic":                    ["AM", "IMS"],
    "stock_warehouse":                 ["AM", "IMS"],
    "demand_120d_daysmart":            ["AM", "IMS"],
    "demand_120d_shopify":             ["AM", "IMS"],
    "demand_120d_hktv":                ["AM", "IMS"],
    "unfulfilled_jit":                 ["AM", "IMS"],
    "upcoming_14d_autoship":           ["AM", "IMS"],
    "competitor_lowest_price":         ["AM", "PS", "IMS"],
    "competitor_source_url":           ["IMS"],
    "competitor_last_reviewed":        ["IMS"],
    "rrp":                             ["PS", "IMS"],
    "sf_express_fee":                  ["AM", "PS", "IMS"],
    "expiration_date":                 ["AM"],
    "bulk_purchase_benefit_available": ["AM"],
    "last_updated_by":                 ["IMS"],
    "last_updated_at":                 ["IMS"],
}

# Filter out columns with no consumer. Warn so we notice.
COLUMNS = []
DROPPED = []
for key, desc in COLUMNS_RAW:
    cons = CONSUMERS_BY_KEY.get(key, [])
    if cons:
        COLUMNS.append((key, desc, cons))
    else:
        DROPPED.append(key)

# ─── 5 example SKUs (real OCR data from current catalogue_items) ───────────
# OCR completeness: 8/8 on the OCR-extracted fields for all 5
EXAMPLES = [
    # All numeric values are pure numbers. Units / context live in the column description.
    # SKU 1: DOXY Doxycycline 100mg Tablets (Medicine, clinic-only) - OCR id 149
    {
        "sku_id":                          "50012149",
        "sku_name":                        "DOXY (Doxycycline) 100mg Tablets — Dechra",
        "brand":                           "Dechra",
        "category":                        "Medicine",
        "sub_category":                    "Antibiotics",
        "species":                         "Dog",
        "storage_rule":                    "clinic_only",
        "status":                          "ACTIVE",
        "hero_sku":                        "FALSE",
        "primary_supplier_id":             "SUP_ALF",
        "supplier_sku_code":               "DO5250",
        "supplier_moq_cached":             "1500",
        "catalogue_cost":                  "680",
        "mbb_pack_cost":                   "645",
        "mbb_min_qty":                     "3",
        "mbb_tier_structure":              "[{\"qty\":3,\"price\":645}]",
        "cost_source":                     "catalogue",
        "cost_last_reviewed_at":           "",
        "minimum_operating_unit":          "1",
        "uom_sell_unit":                   "tablet",
        "minimum_purchase_qty":            "150",
        "units_per_pack":                  "150",
        "bulk_purchase_benefit_available": "TRUE",
        "last_updated_by":                 "ocr-pipeline",
        "last_updated_at":                 "2026-05-31",
    },
    # SKU 2: Entyce capromorelin (Medicine, clinic-only) - OCR id 170
    {
        "sku_id":                          "50012170",
        "sku_name":                        "Entyce® (capromorelin oral solution) 30 mg/mL for dogs — Elanco",
        "brand":                           "Elanco",
        "category":                        "Medicine",
        "sub_category":                    "Appetite Stimulant",
        "species":                         "Dog",
        "storage_rule":                    "clinic_only",
        "status":                          "ACTIVE",
        "hero_sku":                        "FALSE",
        "primary_supplier_id":             "SUP_ALF",
        "supplier_sku_code":               "EN7502",
        "supplier_moq_cached":             "1500",
        "catalogue_cost":                  "1390",
        "mbb_pack_cost":                   "1100",
        "mbb_min_qty":                     "6",
        "mbb_tier_structure":              "[{\"qty\":6,\"price\":1100}]",
        "cost_source":                     "catalogue",
        "cost_last_reviewed_at":           "",
        "minimum_operating_unit":          "1",
        "uom_sell_unit":                   "mL",
        "minimum_purchase_qty":            "30",
        "units_per_pack":                  "30",
        "bulk_purchase_benefit_available": "TRUE",
        "last_updated_by":                 "ocr-pipeline",
        "last_updated_at":                 "2026-05-31",
    },
    # SKU 3: ALOVEEN Shampoo (Pet Hygiene) - OCR id 211
    {
        "sku_id":                          "50012211",
        "sku_name":                        "ALOVEEN Shampoo 250ml — Dermcare",
        "brand":                           "Dermcare",
        "category":                        "Pet Hygiene",
        "sub_category":                    "Shampoo",
        "species":                         "Both",
        "storage_rule":                    "any",
        "status":                          "ACTIVE",
        "hero_sku":                        "FALSE",
        "primary_supplier_id":             "SUP_ALF",
        "supplier_sku_code":               "ALO250",
        "supplier_moq_cached":             "1500",
        "catalogue_cost":                  "58",
        "mbb_pack_cost":                   "54",
        "mbb_min_qty":                     "40",
        "mbb_tier_structure":              "[{\"qty\":10,\"price\":56},{\"qty\":40,\"price\":54}]",
        "cost_source":                     "catalogue",
        "cost_last_reviewed_at":           "",
        "minimum_operating_unit":          "1",
        "uom_sell_unit":                   "bottle",
        "minimum_purchase_qty":            "1",
        "units_per_pack":                  "1",
        "bulk_purchase_benefit_available": "TRUE",
        "last_updated_by":                 "ocr-pipeline",
        "last_updated_at":                 "2026-05-31",
    },
    # SKU 4: Lubrithal Eye Gel (Pet Hygiene) - OCR id 315 - MATCHED to master 59999992
    {
        "sku_id":                          "59999992",
        "sku_name":                        "Dechra - Lubrithal Eye Gel for Cats & Dogs - 15G",
        "brand":                           "Dechra",
        "category":                        "Pet Hygiene",
        "sub_category":                    "Eye Care",
        "species":                         "Both",
        "storage_rule":                    "any",
        "status":                          "ACTIVE",
        "hero_sku":                        "FALSE",
        "primary_supplier_id":             "SUP_ALF",
        "supplier_sku_code":               "D031329",
        "supplier_moq_cached":             "1500",
        "catalogue_cost":                  "112",
        "mbb_pack_cost":                   "107",
        "mbb_min_qty":                     "12",
        "mbb_tier_structure":              "[{\"qty\":12,\"price\":107}]",
        "cost_source":                     "catalogue",
        "cost_last_reviewed_at":           "2026-05-18",
        "minimum_operating_unit":          "1",
        "uom_sell_unit":                   "tube",
        "minimum_purchase_qty":            "1",
        "units_per_pack":                  "1",
        "weight_grams":                    "10",
        "bulk_purchase_benefit_available": "TRUE",
        "last_updated_by":                 "chris@algogroup.io",
        "last_updated_at":                 "2026-05-18",
    },
    # SKU 5: Oral Cleansing Wipes (Pet Hygiene - Dental) - OCR id 232
    {
        "sku_id":                          "50012232",
        "sku_name":                        "Oral Cleansing Wipes — MAXI/GUARD (Addison)",
        "brand":                           "Addison",
        "category":                        "Pet Hygiene",
        "sub_category":                    "Dental",
        "species":                         "Both",
        "storage_rule":                    "any",
        "status":                          "ACTIVE",
        "hero_sku":                        "FALSE",
        "primary_supplier_id":             "SUP_ALF",
        "supplier_sku_code":               "AD011Wipe",
        "supplier_moq_cached":             "1500",
        "catalogue_cost":                  "92",
        "mbb_pack_cost":                   "87",
        "mbb_min_qty":                     "12",
        "mbb_tier_structure":              "[{\"qty\":6,\"price\":90},{\"qty\":12,\"price\":87}]",
        "cost_source":                     "catalogue",
        "cost_last_reviewed_at":           "",
        "minimum_operating_unit":          "1",
        "uom_sell_unit":                   "pot",
        "minimum_purchase_qty":            "1",
        "units_per_pack":                  "1",
        "bulk_purchase_benefit_available": "TRUE",
        "last_updated_by":                 "ocr-pipeline",
        "last_updated_at":                 "2026-05-31",
    },
]

# ─── Business rules: when a field is N/A for a given SKU ──────────────────
# Currently DISABLED per Chris (2026-05-31): "just mark all as blanket incomplete
# for now. easier that way. then we refine the rules later."
#
# The rule mechanism is preserved here so we can switch it back on column-by-column
# as business rules are agreed. Each rule returns a string ("No (rule: …)" or
# "N/A (rule: …)") or None.
#
# When refining: keep rules narrow and named. Always cite the rule in the value
# so the team can see WHY a cell is N/A rather than guess.
RULES_ENABLED = False

def na_rule(col_key, sku):
    if not RULES_ENABLED:
        return None
    # --- example rules to re-enable later ---
    # cat, storage = sku.get("category", ""), sku.get("storage_rule", "")
    # if cat == "Medicine" and storage == "clinic_only":
    #     if col_key == "listed_on_hktv": return "No (rule: clinic-only Medicine)"
    #     ...
    return None


def cell_value(col_key, sku):
    # Known value
    if col_key in sku:
        return sku[col_key]
    # Business rule
    rule = na_rule(col_key, sku)
    if rule:
        return rule
    # Otherwise: required but not yet collected
    return "incomplete"


VERSION = "v6"
LIVE_PAGE_URL = "http://localhost:3001/ssot-spec  (or https://rosetta-ims.vercel.app/ssot-spec)"

# ─── Read Me content (rendered on tab 1 of the XLSX) ──────────────────────
READ_ME_ROWS = [
    [f"SSOT SPEC — SKU MASTER ({VERSION})"],
    [f"Full visual context → {LIVE_PAGE_URL}"],
    ["Last updated: 2026-05-31"],
    ["Owner: Chris Fung"],
    [""],
    ["WHAT IS THIS SHEET"],
    ["The SKU Operational Database tab is the operational database template. Each column is one piece of data we track per SKU."],
    ["Filter or sort by any row 1-3 attribute to navigate. The 5 example SKUs (rows 4-8) show what real data looks like."],
    [""],
    ["OWNER BADGES (row 2 of SKU Operational Database tab)"],
    ["🔵 TECH",         "Tech team (Desmond Brown / Austin) pulls this from a source system pipeline. Don't fill manually."],
    ["🟡 OCR+HUMAN",    "Rosetta IMS OCR extracts; BizOps verifies in /data-review. Cost columns 3-way matched long-term."],
    ["🟢 MANUAL",       "BizOps types it in. The row-2 cell tells you where to look or what to choose."],
    ["⚪ AUTO",         "Rosetta IMS generates this itself (primary keys, timestamps, cached lookups). Ignore."],
    [""],
    ["CONSUMER BADGES (row 3 of SKU Operational Database tab)"],
    ["🟧 AM",           "Approval Matrix — existing-SKU PO approval decisions"],
    ["🟦 PS",           "Product Selection — new-SKU sourcing & listing decisions"],
    ["🟪 IMS",          "Rosetta IMS — read by the IMS itself for UI, audit, lookup"],
    ["🟩 → Shopify",    "Pushed BACK to Shopify (e.g. minimum qty, multiples settings)"],
    ["🟨 → DaySmart",   "Pushed BACK to DaySmart (e.g. cost overrides)"],
    [""],
    ["ROW STRUCTURE OF THE SKU OPERATIONAL DATABASE TAB"],
    ["Row 1", "Column header (sku_id, sku_name, brand, …)"],
    ["Row 2", "Source / how-to-fill description with owner badge"],
    ["Row 3", "Consumers — who reads this column (one per line)"],
    ["Row 4-8", "5 real example SKUs (pulled from current Rosetta IMS OCR scan, where data is complete)"],
    [""],
    ["CELL VALUE CONVENTIONS"],
    ["PURE TYPES ONLY", "Every value must be an atomic, machine-readable type: NUMBER, INTEGER, ISO DATE, ENUM string, or JSON. NEVER mix number + unit + parenthetical (e.g. '150 tabs (1 bottle)' is WRONG — write '150', let uom_sell_unit + description carry the unit)."],
    ["why",             "Sam wires these cells directly into Approval Matrix formulas. Logic Layer arithmetic breaks the moment a string is in a numeric cell. Cloud migration will enforce types — better to comply now."],
    ["<value>",         "Known value (from OCR, master list, or computed)"],
    ["incomplete",      "Required but not yet collected (OCR didn't extract, tech pipeline not built, or human hasn't typed)"],
    ["N/A (rule: …)",   "Not applicable for this SKU per a named business rule (rules are currently disabled — all missing = incomplete)"],
    [""],
    ["ACCURACY LADDER (where the cost columns climb to 100%)"],
    ["Step A", "OCR machine extraction → /catalogues. Not trusted alone."],
    ["Step B", "OCR + Human review → /data-review. ~80% accurate COMBINED (humans err too)."],
    ["Step C", "3-way matching vs accounts (PO ↔ delivery note ↔ invoice). Closes remaining ~20%."],
    [""],
    ["HOW TO REGENERATE THIS SHEET"],
    ["1.", "Edit backend/scripts/gen_ssot_spec_csv.py (COLUMNS, CONSUMERS_BY_KEY, EXAMPLES)"],
    ["2.", "Run: python backend/scripts/gen_ssot_spec_csv.py"],
    ["3.", "Upload the resulting docs/ssot-spec.xlsx to Google Drive (creates a fresh sheet)"],
    ["4.", "Update SHEET_URL in frontend/src/app/ssot-spec/page.tsx"],
    [""],
    ["TASK FOR SAM & SEPH — PLUG-IN TEST"],
    ["Assume", "The legacy SSOT Database sheet is DELETED Monday. The only data input going forward is supplier catalogues processed by Rosetta IMS OCR. Assume that OCR + Human review + 3-way matching means every column on the SKU Operational Database tab will reach 100% populated."],
    ["Your job", "Walk through every column in your existing tab (Sam: Approval Matrix; Seph: Product Selection). For each column, can you point at the SKU Operational Database tab and say 'yes, that's where I'd pull from'? List the gaps."],
    ["Deliverable", "A 3-column table: column_in_my_tab | maps_to_v8_column | gap (🟢 GREEN / 🟡 AMBER / 🔴 RED). Reply by Friday."],
    ["🟢 GREEN", "direct map — SKU Operational Database has the column you need"],
    ["🟡 AMBER", "exists in SKU Operational Database but semantics unclear or partially covered"],
    ["🔴 RED",   "missing from SKU Operational Database; required for your decision logic"],
    ["", "Red rows become the v9 backlog."],
    ["Best practice", "When editing the SKU Operational Database tab: duplicate the tab, highlight your changes, and share for alignment. Claude pulls from this sheet to update Rosetta IMS."],
]

# ─── Plug-in Test content (rendered on tab 3 of the XLSX) ─────────────────
# A red/amber/green template for Sam (AM) and Seph (PS) to validate whether
# v7 covers every column they currently use. Pre-filled with examples.
PLUG_IN_TEST_HEADER = [
    ["PLUG-IN TEST — fill this in"],
    ["For Sam (Approval Matrix) and Seph (Product Selection)."],
    ["Assume the legacy SSOT Database is deleted Monday. Walk through every column in your existing tab. For each, mark whether the v7 SKU Operational Database tab covers it."],
    [""],
    ["🟢 GREEN  = direct map, v7 has the column you need"],
    ["🟡 AMBER  = exists in v7 but semantics unclear or partially covered"],
    ["🔴 RED    = missing from v7; required for your decision logic"],
    [""],
    ["Deliverable: complete this tab and send back. Red rows become the v8 backlog."],
    [""],
]

PLUG_IN_TEST_COLUMNS = ["workstream", "column_in_my_tab", "maps_to_v7_column", "gap", "notes"]

# Pre-filled illustrative examples (mix of green/amber/red across AM and PS)
PLUG_IN_TEST_EXAMPLES = [
    # workstream, column_in_my_tab, maps_to_v7_column, gap, notes
    ["── EXAMPLES (read these to learn the format) ──", "", "", "", ""],
    ["AM",  "Wholesale Cost basic",               "basic_unit_cost",                  "🟢 GREEN", "Direct map."],
    ["AM",  "Wholesale Cost MBB",                 "mbb_cost_per_unit",                "🟢 GREEN", "Direct map."],
    ["AM",  "Qty/Box",                            "units_per_pack",                   "🟡 AMBER", "v7 covers it but semantics confused for liquids — stores volume per bottle, not units per case."],
    ["AM",  "Selling Price From DaySmart",        "sell_price_clinic",                "🟢 GREEN", "Tech pull pipeline to build."],
    ["AM",  "Competitor Price (HKTV)",            "competitor_lowest_price",          "🟡 AMBER", "v7 has 'lowest' but doesn't say WHICH competitor. Need competitor_name field too."],
    ["AM",  "Required GP%",                       "(formula in Logic Layer)",         "🟢 GREEN", "Comes from category lookup, not SSOT. By design."],
    ["AM",  "Reorder point / Safety stock",       "(MISSING from v7)",                "🔴 RED",   "Needed for WOC trigger logic. Should be added or derived."],
    ["AM",  "Lead time (days)",                   "(MISSING from v7)",                "🔴 RED",   "Implicit in Suppliers.cutoff_day + next_delivery_day but not explicit. Needed for Post-Purchase WOC."],
    ["PS",  "Vetopia Selling Price",              "competitor_lowest_price",          "🟡 AMBER", "v7 aggregates competitors. Vetopia is a specific source — may want per-competitor pricing."],
    ["PS",  "Supplier MOQ",                       "supplier_moq_cached",              "🟢 GREEN", "Cached from Suppliers table."],
    ["PS",  "Risk acceptance (we_take_risk etc.)","risk_acceptance",                  "🟢 GREEN", "NEW per 29 May call. Direct map."],
    ["PS",  "Sub-category granularity",           "sub_category",                     "🟢 GREEN", "NEW per 29 May call. Direct map."],
    ["PS",  "Competitor selling volume (HKTV)",   "(MISSING from v7)",                "🔴 RED",   "Market-demand signal for List/Not List. Need to add."],
    ["PS",  "Controlled Price flag",              "(MISSING from v7)",                "🔴 RED",   "Flag for price-regulated SKUs. Existing master has this; v7 dropped it."],
    ["── YOUR ROWS GO HERE (add as many as you need) ──", "", "", "", ""],
    ["AM",  "",                                   "",                                 "",         ""],
    ["AM",  "",                                   "",                                 "",         ""],
    ["AM",  "",                                   "",                                 "",         ""],
    ["AM",  "",                                   "",                                 "",         ""],
    ["AM",  "",                                   "",                                 "",         ""],
    ["AM",  "",                                   "",                                 "",         ""],
    ["AM",  "",                                   "",                                 "",         ""],
    ["AM",  "",                                   "",                                 "",         ""],
    ["AM",  "",                                   "",                                 "",         ""],
    ["AM",  "",                                   "",                                 "",         ""],
    ["PS",  "",                                   "",                                 "",         ""],
    ["PS",  "",                                   "",                                 "",         ""],
    ["PS",  "",                                   "",                                 "",         ""],
    ["PS",  "",                                   "",                                 "",         ""],
    ["PS",  "",                                   "",                                 "",         ""],
    ["PS",  "",                                   "",                                 "",         ""],
    ["PS",  "",                                   "",                                 "",         ""],
    ["PS",  "",                                   "",                                 "",         ""],
    ["PS",  "",                                   "",                                 "",         ""],
    ["PS",  "",                                   "",                                 "",         ""],
]


def write_csv(out_path):
    with open(out_path, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow([c[0] for c in COLUMNS])
        w.writerow([c[1] for c in COLUMNS])
        w.writerow([consumers_text(c[2]) for c in COLUMNS])
        for sku in EXAMPLES:
            w.writerow([cell_value(c[0], sku) for c in COLUMNS])


def write_xlsx(out_path):
    if not OPENPYXL_OK:
        print("  (skipped XLSX — openpyxl not installed; pip install openpyxl)")
        return False
    wb = openpyxl.Workbook()

    # Tab 1: Read Me
    ws_readme = wb.active
    ws_readme.title = "Read Me"
    ws_readme.column_dimensions["A"].width = 22
    ws_readme.column_dimensions["B"].width = 110
    bold = Font(bold=True, size=11)
    header_fill = PatternFill("solid", fgColor="1E293B")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    wrap = Alignment(wrap_text=True, vertical="top")
    section_keys = {"WHAT IS THIS SHEET","OWNER BADGES (row 2 of SKU Operational Database tab)",
                    "CONSUMER BADGES (row 3 of SKU Operational Database tab)","ROW STRUCTURE OF THE SKU OPERATIONAL DATABASE TAB",
                    "CELL VALUE CONVENTIONS","ACCURACY LADDER (where the cost columns climb to 100%)",
                    "HOW TO REGENERATE THIS SHEET","TASK FOR SAM & SEPH — PLUG-IN TEST"}
    for r, row in enumerate(READ_ME_ROWS, start=1):
        for c, val in enumerate(row, start=1):
            cell = ws_readme.cell(row=r, column=c, value=val)
            cell.alignment = wrap
            if r == 1:
                cell.font = Font(bold=True, size=16, color="0F172A")
            elif r in (2, 3, 4):
                cell.font = Font(color="64748B", italic=True)
            elif len(row) == 1 and val in section_keys:
                cell.font = header_font
                cell.fill = header_fill
            elif len(row) >= 2 and c == 1:
                cell.font = bold

    # Tab 2: SKU Operational Database (renamed from "SKU Master" per v7 review on 2026-06-01)
    ws = wb.create_sheet("SKU Operational Database")
    ws.append([c[0] for c in COLUMNS])
    ws.append([c[1] for c in COLUMNS])
    ws.append([consumers_text(c[2]) for c in COLUMNS])

    # Resolve column letters for formula references
    def col_idx_of(key):
        for i, c in enumerate(COLUMNS, start=1):
            if c[0] == key:
                return i
        return None
    cat_col_letter      = openpyxl.utils.get_column_letter(col_idx_of('catalogue_cost'))
    upp_col_letter      = openpyxl.utils.get_column_letter(col_idx_of('units_per_pack'))
    mbb_pack_col_letter = openpyxl.utils.get_column_letter(col_idx_of('mbb_pack_cost'))

    for r_offset, sku in enumerate(EXAMPLES):
        row_num = r_offset + 4  # row 1=header, 2=desc, 3=consumer, 4+=data
        for c_idx, (key, _desc, _cons) in enumerate(COLUMNS, start=1):
            if key == 'basic_unit_cost':
                # FORMULA: =catalogue_cost / units_per_pack
                val = f"={cat_col_letter}{row_num}/{upp_col_letter}{row_num}"
            elif key == 'mbb_cost_per_unit':
                # FORMULA: =mbb_pack_cost / units_per_pack
                val = f"={mbb_pack_col_letter}{row_num}/{upp_col_letter}{row_num}"
            else:
                val = cell_value(key, sku)
            ws.cell(row=row_num, column=c_idx, value=val)

    # Style row 1 (headers) and row 2/3 (metadata)
    header_row_fill = PatternFill("solid", fgColor="6366F1")
    header_row_font = Font(bold=True, color="FFFFFF", size=11)
    meta_fill = PatternFill("solid", fgColor="F1F5F9")
    consumer_fill = PatternFill("solid", fgColor="FFFBEB")
    wrap_top = Alignment(wrap_text=True, vertical="top")
    for col_idx in range(1, len(COLUMNS) + 1):
        c1 = ws.cell(row=1, column=col_idx)
        c1.fill = header_row_fill
        c1.font = header_row_font
        c1.alignment = Alignment(vertical="center")
        c2 = ws.cell(row=2, column=col_idx)
        c2.fill = meta_fill
        c2.alignment = wrap_top
        c2.font = Font(size=9, color="475569")
        c3 = ws.cell(row=3, column=col_idx)
        c3.fill = consumer_fill
        c3.alignment = wrap_top
        c3.font = Font(size=9, color="78350F")

    ws.freeze_panes = "A4"  # Freeze rows 1-3 (header/desc/consumer)
    for col_idx in range(1, len(COLUMNS) + 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 22
    ws.row_dimensions[2].height = 60
    ws.row_dimensions[3].height = 70

    wb.save(out_path)
    return True

# (Plug-in Test tab removed; instructions are in the Read Me tab.)
def _DEAD_legacy_plug_in_test_code(wb):
    # ─── Tab 3: Plug-in Test ───
    wp = wb.create_sheet("Plug-in Test")
    wp.column_dimensions["A"].width = 14   # workstream
    wp.column_dimensions["B"].width = 38   # column_in_my_tab
    wp.column_dimensions["C"].width = 38   # maps_to_v7_column
    wp.column_dimensions["D"].width = 12   # gap
    wp.column_dimensions["E"].width = 70   # notes

    # Header instruction rows
    cur = 1
    title_font = Font(bold=True, size=16, color="0F172A")
    italic_grey = Font(color="64748B", italic=True, size=11)
    green_font = Font(color="166534", bold=True, size=11)
    amber_font = Font(color="92400E", bold=True, size=11)
    red_font = Font(color="991B1B", bold=True, size=11)
    for row in PLUG_IN_TEST_HEADER:
        for c, val in enumerate(row, start=1):
            cell = wp.cell(row=cur, column=c, value=val)
            if cur == 1:
                cell.font = title_font
            elif "GREEN" in str(val):
                cell.font = green_font
            elif "AMBER" in str(val):
                cell.font = amber_font
            elif "RED" in str(val):
                cell.font = red_font
            else:
                cell.font = italic_grey
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        cur += 1

    # Header row for the table
    header_fill_pt = PatternFill("solid", fgColor="0F172A")
    header_font_pt = Font(bold=True, color="FFFFFF", size=11)
    for c, val in enumerate(PLUG_IN_TEST_COLUMNS, start=1):
        cell = wp.cell(row=cur, column=c, value=val)
        cell.fill = header_fill_pt
        cell.font = header_font_pt
        cell.alignment = Alignment(vertical="center")
    cur += 1

    # Example + blank rows
    section_fill = PatternFill("solid", fgColor="F1F5F9")
    section_font = Font(italic=True, color="475569", bold=True, size=10)
    am_fill = PatternFill("solid", fgColor="FFEDD5")
    ps_fill = PatternFill("solid", fgColor="DBEAFE")
    green_fill = PatternFill("solid", fgColor="DCFCE7")
    amber_fill = PatternFill("solid", fgColor="FEF3C7")
    red_fill = PatternFill("solid", fgColor="FEE2E2")

    for row in PLUG_IN_TEST_EXAMPLES:
        is_section = row[0].startswith("──")
        ws_val = row[0]
        for c, val in enumerate(row, start=1):
            cell = wp.cell(row=cur, column=c, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if is_section:
                cell.fill = section_fill
                cell.font = section_font
            else:
                if c == 1:  # workstream chip
                    if ws_val == "AM":
                        cell.fill = am_fill
                        cell.font = Font(bold=True, color="7C2D12", size=11)
                    elif ws_val == "PS":
                        cell.fill = ps_fill
                        cell.font = Font(bold=True, color="1E40AF", size=11)
                if c == 4:  # gap column colour
                    if "GREEN" in str(val):
                        cell.fill = green_fill
                        cell.font = Font(bold=True, color="166534")
                    elif "AMBER" in str(val):
                        cell.fill = amber_fill
                        cell.font = Font(bold=True, color="92400E")
                    elif "RED" in str(val):
                        cell.fill = red_fill
                        cell.font = Font(bold=True, color="991B1B")
        cur += 1

    wp.freeze_panes = f"A{len(PLUG_IN_TEST_HEADER) + 2}"  # Freeze header rows
    return True


def main():
    docs_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "docs"))
    csv_path  = os.path.join(docs_dir, "ssot-spec.csv")
    xlsx_path = os.path.join(docs_dir, "ssot-spec.xlsx")

    write_csv(csv_path)
    print(f"Wrote {csv_path}")
    print(f"  {len(COLUMNS)} columns × {len(EXAMPLES) + 3} rows (header + description + consumers + {len(EXAMPLES)} SKUs)")

    ok = write_xlsx(xlsx_path)
    if ok:
        print(f"Wrote {xlsx_path}  (2 tabs: 'Read Me' + 'SKU Operational Database')")

    if DROPPED:
        print(f"  DROPPED {len(DROPPED)} column(s) with no downstream consumer: {DROPPED}")
    else:
        print(f"  All {len(COLUMNS)} columns have at least one consumer.")


if __name__ == "__main__":
    main()
