"""
AI-assisted catalogue extraction service.

Uses Claude claude-haiku-4-5-20251001 (fast, cheap) for PDF/image extraction.
Falls back to rule-based parsing for Excel/CSV files.

Set ANTHROPIC_API_KEY in the environment to enable AI extraction.
Without the key, extraction returns a stub result with confidence_score=0
so the review queue still works — items just need manual field entry.
"""
import os
import io
import json
import re
import base64
from typing import Optional
import openpyxl
import pypdf

ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")

# Haiku 4.5 supports a large output budget. Dense catalogues produce long JSON
# arrays — too small a cap truncates mid-object and the whole parse fails.
MAX_TOKENS = 8192


def _loads_json_array(raw: str) -> list[dict]:
    """Parse a JSON array of objects, tolerating markdown fences and truncation.

    Claude occasionally hits the output-token cap and returns a JSON array cut
    off mid-object. Rather than lose the entire chunk to a JSONDecodeError, we
    salvage every complete top-level {...} object the model did manage to emit.
    """
    raw = raw.strip()
    raw = re.sub(r'^```(?:json)?\s*', '', raw)
    raw = re.sub(r'\s*```$', '', raw)
    try:
        data = json.loads(raw)
        return data if isinstance(data, list) else []
    except json.JSONDecodeError:
        pass
    # Salvage: walk the array and collect each balanced {...} object.
    start = raw.find('[')
    if start == -1:
        return []
    objs: list[dict] = []
    depth = 0
    obj_start: int | None = None
    in_str = esc = False
    for i in range(start + 1, len(raw)):
        c = raw[i]
        if in_str:
            if esc:            esc = False
            elif c == '\\':    esc = True
            elif c == '"':     in_str = False
            continue
        if c == '"':
            in_str = True
        elif c == '{':
            if depth == 0:
                obj_start = i
            depth += 1
        elif c == '}':
            depth -= 1
            if depth == 0 and obj_start is not None:
                try:
                    objs.append(json.loads(raw[obj_start:i + 1]))
                except json.JSONDecodeError:
                    pass
                obj_start = None
    return objs

EXTRACTION_PROMPT = """You are extracting product data from a supplier price list for a Hong Kong veterinary/pet supply company.

Extract ALL product entries from this document. For each product return a JSON object with these fields:
- description: product name/description (string)
- brand: the product's brand / manufacturer if identifiable from the line (string or null) — e.g. "Royal Canin", "Dechra", "Hill's"
- species: the target animal — one of "dog", "cat", "both", "other" if stated/clear from the product, else null
- supplier_sku: supplier's own product code/SKU (string or null)
- barcode: EAN/UPC barcode if present (string or null)
- cost_price: unit cost price in HKD as a number (float or null) — strip currency symbols
- rrp: recommended retail price / RRP / SRP in HKD if the catalogue prints one (float or null) — distinct from cost
- uom: the SELL unit of measurement — what ONE sellable unit is, e.g. "tablet", "capsule", "bottle", "ml", "kg", "each" (string or null)
- min_sellable_qty: the SMALLEST quantity sold in a single sale, in `uom` units — almost always 1 (you sell 1 tablet/capsule/sachet at a time). Only >1 if the product can ONLY be sold as a fixed bundle (e.g. dispensed only as a sealed strip of 10). Default 1 for any discrete sellable unit (integer or null)
- units_per_pack: how many sell-units come in one purchasable pack/box/strip/bottle — i.e. the whole pack you must buy. E.g. a box of 100 tablets → uom="tablet", min_sellable_qty=1, units_per_pack=100 (integer or null)
- pack_size: the raw pack-size text as printed (string or null) — e.g. "100 tabs/box", "250ml", "strip of 14"
- variant: the size / volume / weight / flavour that distinguishes THIS item from its sibling variations of the same base product — e.g. "15ml", "2kg", "Chicken", "Large". null if the product has no size/flavour variants.
- weight_grams: net weight of ONE sell-unit in grams if stated (number or null). ALWAYS convert to grams using the unit ACTUALLY printed — kg→g (e.g. "2kg" → 2000, "85g" → 85) and pounds→g at 1 lb = 453.592 g (e.g. "3 LBs" → 1361, "5lb" → 2268). Do NOT treat a pound value as kilograms.
- min_purchase_qty: supplier MINIMUM ORDER QUANTITY (MOQ) — how many PACKS you must order at once, if stated. This is DISTINCT from units_per_pack/min_sellable_qty; leave null unless the supplier states an order minimum (integer or null)
- order_increment_qty: the ORDER MULTIPLE / carton — if the price list has an "order multiple" / "order units" / "carton of N" column or note, the number of sell-units you must order in multiples of. DISTINCT from units_per_pack (the cost-dividing pack size). Leave null unless stated (integer or null)
- bulk_buy_tiers: human-readable bulk discount tiers if present, e.g. "6 bots @ 92; 10 bots @ 88" or "9+1 boxes" (string or null)
- bulk_tiers: STRUCTURED version of the discount tiers — array of {"min_qty": <int>, "unit_cost": <float>} per tier, or null if none
- max_bulk_buy_cost: deepest-discount per-unit cost across all bulk tiers (float or null) — the lowest quoted price, NOT the basic. For "buy N get M free" deals without a discounted unit price, set to null.
- max_bulk_buy_min_qty: minimum order quantity required to qualify for max_bulk_buy_cost (integer or null)
- confidence: overall confidence for this row as 0.0-1.0 (float)

Return ONLY a JSON array of these objects, no explanation. If a field is missing or unclear, use null.
Example: [{"description":"Apoquel 16mg Tab x100","brand":"Zoetis","species":"dog","supplier_sku":"APQ16","barcode":"5415198219506","cost_price":1850.00,"rrp":2400.00,"uom":"tablet","min_sellable_qty":1,"units_per_pack":100,"pack_size":"100 tabs/bottle","variant":null,"weight_grams":null,"min_purchase_qty":null,"order_increment_qty":null,"bulk_buy_tiers":"5 bots @ 1750; 10 bots @ 1690","bulk_tiers":[{"min_qty":5,"unit_cost":1750.00},{"min_qty":10,"unit_cost":1690.00}],"max_bulk_buy_cost":1690.00,"max_bulk_buy_min_qty":10,"confidence":0.95}]
Variations example — the single line "Vetericyn Plus Wound Spray  15ml $50 / 45ml $120 / 60ml $150" becomes THREE objects:
[{"description":"Vetericyn Plus Wound & Skin Care Spray 15ml","brand":"Vetericyn","variant":"15ml","cost_price":50.00,"uom":"ml","pack_size":"15ml","confidence":0.9},
 {"description":"Vetericyn Plus Wound & Skin Care Spray 45ml","brand":"Vetericyn","variant":"45ml","cost_price":120.00,"uom":"ml","pack_size":"45ml","confidence":0.9},
 {"description":"Vetericyn Plus Wound & Skin Care Spray 60ml","brand":"Vetericyn","variant":"60ml","cost_price":150.00,"uom":"ml","pack_size":"60ml","confidence":0.9}]

Important:
- Extract ALL rows, even if some fields are missing
- Do NOT skip rows — incomplete rows are useful for the review queue
- VARIATIONS — THIS IS CRITICAL: when one product is offered in MULTIPLE sizes / volumes /
  weights / flavours (e.g. 15ml, 45ml, 60ml; or 2kg / 5kg / 12kg; or Chicken / Beef), output
  a SEPARATE object for EACH variation — NEVER merge them into one row. Each variation:
    • gets its own `variant` (e.g. "15ml") and its `description` MUST include that size/flavour
      so each is a distinct product (e.g. "Vetericyn Wound Spray 15ml", "... 45ml", "... 60ml");
    • carries its own cost_price, rrp, supplier_sku, barcode, pack_size and weight where the
      catalogue lists per-variation values; if a value is shown once for the whole group,
      repeat it on each variation.
  A single line that lists several size→price pairs (in columns OR inline, e.g.
  "15ml $50 / 45ml $120 / 60ml $150", or a row with columns "15ml | 45ml | 60ml" each with a
  price) represents MULTIPLE products — emit one object per size, not one combined row.
- If this appears to be a clinical consumable (surgical drapes, catheters, syringes) rather than a retail product, still include it — the human reviewer will filter
- Prices in HKD: strip "$", "HKD", "HK$" prefixes, strip commas
- Pack size: "100 tabs", "box/100", "strip of 14" all indicate units_per_pack — extract the number into units_per_pack AND keep the printed text in pack_size
- weight: for food/litter the net weight is usually on the line (e.g. "2kg", "85g pouch", "3 LBs") — convert to grams using the printed unit (kg→g; pounds at 1 lb = 453.592 g, e.g. "3 LBs" → 1361). For tablets/liquids leave null unless a per-unit weight is printed
- cost_price = STANDARD (smallest-quantity) unit cost. max_bulk_buy_cost = LOWEST tier price. If the only "deal" is buy-N-get-M-free with no quoted discount price, leave max_bulk_buy_cost null but include the deal in bulk_buy_tiers.
- Only fill brand/species when the line makes it clear — do NOT guess.
"""


def _guided_prompt(contract=None) -> str:
    """Base extraction prompt, plus the supplier's contract guidance when one applies (DC-2)."""
    return EXTRACTION_PROMPT + (contract.prompt_section() if contract is not None else "")


def _call_claude_vision(image_b64: str, media_type: str, contract=None) -> list[dict]:
    """Call Claude claude-haiku-4-5-20251001 to extract product data from an image or PDF.
    PDFs MUST go in a `document` block — the `image` block only accepts
    image/jpeg|png|gif|webp, so sending a PDF there returns a 400."""
    import anthropic
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
    block_type = "document" if media_type == "application/pdf" else "image"
    msg = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=MAX_TOKENS,
        messages=[{
            "role": "user",
            "content": [
                {
                    "type": block_type,
                    "source": {"type": "base64", "media_type": media_type, "data": image_b64},
                },
                {"type": "text", "text": _guided_prompt(contract)},
            ],
        }],
    )
    return _loads_json_array(msg.content[0].text)


def _call_claude_text(text: str, contract=None) -> list[dict]:
    """Call Claude Haiku to extract product data from text content.
    Chunks large documents so nothing is silently truncated.
    """
    import anthropic
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

    CHUNK = 8000    # chars per chunk — smaller so the JSON output fits MAX_TOKENS
    OVERLAP = 500   # overlap so line items don't get split at chunk boundary

    chunks = []
    start = 0
    while start < len(text):
        end = min(start + CHUNK, len(text))
        # Try to split on a newline so we don't cut mid-row
        if end < len(text):
            newline = text.rfind('\n', start + CHUNK - OVERLAP, end)
            if newline != -1:
                end = newline
        chunks.append(text[start:end])
        start = end - OVERLAP if end < len(text) else end

    all_items: list[dict] = []
    seen_skus: set[str] = set()
    first_error: str | None = None

    for chunk in chunks:
        try:
            msg = client.messages.create(
                model="claude-haiku-4-5-20251001",
                max_tokens=MAX_TOKENS,
                messages=[{
                    "role": "user",
                    "content": f"{_guided_prompt(contract)}\n\nDocument text:\n{chunk}",
                }],
            )
            items = _loads_json_array(msg.content[0].text)
            if not items:
                continue
            for item in items:
                sku = item.get("supplier_sku") or ""
                desc = item.get("description") or ""
                dedup_key = (sku.lower(), desc.lower()[:40])
                if dedup_key not in seen_skus:
                    seen_skus.add(dedup_key)
                    all_items.append(item)
        except Exception as e:
            if first_error is None:
                first_error = f"{type(e).__name__}: {e}"
            continue

    if not all_items and first_error:
        return [{"description": f"[Extraction error] {first_error}", "supplier_sku": None,
                 "barcode": None, "cost_price": None, "uom": None, "units_per_pack": None,
                 "bulk_buy_tiers": None, "confidence": 0.0}]

    return all_items


def _stub_result(filename: str) -> list[dict]:
    """Return a stub when no API key is configured."""
    return [{
        "description": f"[AI extraction disabled — set ANTHROPIC_API_KEY to enable] {filename}",
        "supplier_sku": None,
        "barcode": None,
        "cost_price": None,
        "uom": None,
        "bulk_buy_tiers": None,
        "confidence": 0.0,
        "_stub": True,
    }]


def extract_from_excel(content: bytes, filename: str) -> list[dict]:
    """
    Rule-based extraction from Excel files.
    Reads the first sheet, auto-detects header row, maps common column names.
    """
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # Find header row — first row with 3+ non-empty cells
    header_idx = 0
    for i, row in enumerate(rows[:10]):
        non_empty = sum(1 for c in row if c is not None and str(c).strip())
        if non_empty >= 3:
            header_idx = i
            break

    headers = [str(c).strip().lower() if c else "" for c in rows[header_idx]]

    # Column name heuristics
    def find_col(*keywords):
        for kw in keywords:
            for i, h in enumerate(headers):
                if kw in h:
                    return i
        return None

    col_desc    = find_col('description', 'product', 'name', 'item')
    col_sku     = find_col('sku', 'code', 'ref', 'item no', 'item#', 'part')
    col_barcode = find_col('barcode', 'ean', 'upc', 'gtin')
    col_price   = find_col('price', 'cost', 'rate', 'unit price')
    col_uom     = find_col('uom', 'unit', 'pack', 'size')

    results = []
    for row in rows[header_idx + 1:]:
        if all(c is None or str(c).strip() == "" for c in row):
            continue

        def cell(idx):
            if idx is None or idx >= len(row): return None
            v = row[idx]
            return str(v).strip() if v is not None else None

        desc = cell(col_desc)
        if not desc:
            continue

        raw_price = cell(col_price)
        cost = None
        if raw_price:
            cleaned = re.sub(r'[HKD$,\s]', '', raw_price)
            try:
                cost = float(cleaned)
                if cost <= 0: cost = None
            except (ValueError, TypeError):
                cost = None

        results.append({
            "description":    desc,
            "supplier_sku":   cell(col_sku),
            "barcode":        cell(col_barcode),
            "cost_price":     cost,
            "uom":            cell(col_uom),
            "bulk_buy_tiers": None,
            "confidence":     0.75 if cost else 0.5,
        })

    wb.close()
    return results


def _pdf_text_is_reliable(text: str) -> bool:
    """True if pypdf's extracted text looks usable. Bilingual HK catalogues often use custom
    CID fonts that pypdf decodes to mojibake (glyphs land in random Unicode blocks), silently
    losing the table data — weight variants, SKUs, prices. When too much of the text is such
    'unexpected' characters, treat it as unusable so the caller falls back to native PDF vision."""
    chars = [c for c in text if not c.isspace()]
    if not chars:
        return False

    def expected(ch: str) -> bool:
        o = ord(ch)
        return (
            0x20 <= o <= 0x7E         # printable ASCII
            or 0x00A0 <= o <= 0x00FF  # latin-1 supplement (®, ½, accents)
            or 0x2000 <= o <= 0x206F  # general punctuation (–, •, “”)
            or 0x3000 <= o <= 0x303F  # CJK symbols & punctuation
            or 0x3400 <= o <= 0x4DBF  # CJK Ext-A
            or 0x4E00 <= o <= 0x9FFF  # CJK Unified Ideographs (real Chinese)
            or 0xFF00 <= o <= 0xFFEF  # full/half-width forms
        )

    suspicious = sum(1 for ch in chars if not expected(ch))
    return suspicious / len(chars) < 0.12


def extract_from_pdf(content: bytes, filename: str, contract=None) -> list[dict]:
    """Extract text from PDF, then use Claude to structure it."""
    if not ANTHROPIC_API_KEY:
        return _stub_result(filename)

    reader = pypdf.PdfReader(io.BytesIO(content))
    pages_text = []
    for page in reader.pages:
        t = page.extract_text()
        if t:
            pages_text.append(t)

    full_text = "\n\n--- PAGE BREAK ---\n\n".join(pages_text)
    if not full_text.strip() or not _pdf_text_is_reliable(full_text):
        # No extractable text (scanned PDF) OR garbled text — custom/CID fonts in bilingual
        # HK catalogues decode to mojibake via pypdf and silently drop the table data
        # (weight variants, SKUs, prices). Vision reads the rendered pages natively.
        return extract_from_pdf_as_image(content, filename, contract)

    try:
        items = _call_claude_text(full_text, contract)
        return items if isinstance(items, list) else []
    except Exception as e:
        return [{"description": f"Extraction error: {e}", "supplier_sku": None,
                 "barcode": None, "cost_price": None, "uom": None,
                 "bulk_buy_tiers": None, "confidence": 0.0}]


def extract_from_pdf_as_image(content: bytes, filename: str, contract=None) -> list[dict]:
    """For scanned PDFs (no extractable text): send the PDF to Claude as a `document`
    block so it reads the page images natively. Send the whole file — truncating PDF
    bytes mid-stream corrupts the document and the API rejects it."""
    if not ANTHROPIC_API_KEY:
        return _stub_result(filename)
    b64 = base64.standard_b64encode(content).decode()
    try:
        return _call_claude_vision(b64, "application/pdf", contract)
    except Exception as e:
        return [{"description": f"Vision extraction error: {e}", "supplier_sku": None,
                 "barcode": None, "cost_price": None, "uom": None,
                 "bulk_buy_tiers": None, "confidence": 0.0}]


def extract_from_image(content: bytes, media_type: str, filename: str, contract=None) -> list[dict]:
    """Extract from JPEG/PNG image (e.g. WhatsApp price list photo)."""
    if not ANTHROPIC_API_KEY:
        return _stub_result(filename)
    b64 = base64.standard_b64encode(content).decode()
    try:
        return _call_claude_vision(b64, media_type, contract)
    except Exception as e:
        return [{"description": f"Image extraction error: {e}", "supplier_sku": None,
                 "barcode": None, "cost_price": None, "uom": None,
                 "bulk_buy_tiers": None, "confidence": 0.0}]


DETECTION_PROMPT = """Identify the SUPPLIER (the distributor/company we would place a purchase order with) and the product BRAND(S) in this Hong Kong veterinary / pet-supply price list or catalogue.

Use every signal: the letterhead / company name, document title, header & footer, contact details (phone, email, website domain), and the FILE NAME provided above.

Return ONLY this JSON, no explanation:
{"supplier": "<distributor / supplier company name, or null>", "brands": ["<brand>", ...], "confidence": <0.0-1.0>}

- supplier = the company that SELLS this catalogue to us (the distributor), NOT the manufacturer unless they are the same.
- brands = product brand names featured (e.g. "Royal Canin", "Hill's", "Almo Nature"). Use [] if none are clear.
- confidence = how sure you are about the supplier (0.0-1.0)."""


def _detect_call(content_blocks: list) -> dict:
    import anthropic
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
    msg = client.messages.create(model="claude-haiku-4-5-20251001", max_tokens=400,
                                 messages=[{"role": "user", "content": content_blocks}])
    raw = msg.content[0].text
    i, j = raw.find("{"), raw.rfind("}")
    obj = json.loads(raw[i:j + 1]) if 0 <= i < j else {}
    return {"supplier": obj.get("supplier") or None,
            "brands": [b for b in (obj.get("brands") or []) if b],
            "confidence": float(obj.get("confidence") or 0.0)}


def detect_supplier_brand(content: bytes, filename: str, content_type: str = "") -> dict:
    """Document-level supplier + brand detection for catalogue->supplier matching.

    Reads the letterhead/title/footer/contacts + the filename. Returns
    {supplier, brands, confidence}; degrades to empty on no-API-key or any failure
    (so a detection hiccup never blocks the upload)."""
    empty = {"supplier": None, "brands": [], "confidence": 0.0}
    if not ANTHROPIC_API_KEY:
        return empty
    nl = (filename or "").lower()
    ct = (content_type or "").lower()
    head = f"FILE NAME: {filename}\n\n"
    try:
        if nl.endswith((".jpg", ".jpeg", ".png")) or "image" in ct:
            mt = "image/png" if (nl.endswith(".png") or "png" in ct) else "image/jpeg"
            return _detect_call([
                {"type": "image", "source": {"type": "base64", "media_type": mt,
                                             "data": base64.standard_b64encode(content).decode()}},
                {"type": "text", "text": head + DETECTION_PROMPT}])
        if nl.endswith(".pdf") or "pdf" in ct:
            return _detect_call([
                {"type": "document", "source": {"type": "base64", "media_type": "application/pdf",
                                                "data": base64.standard_b64encode(content).decode()}},
                {"type": "text", "text": head + DETECTION_PROMPT}])
        if nl.endswith((".xlsx", ".xls")) or "spreadsheet" in ct or "excel" in ct:
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            lines = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 20:
                    break
                lines.append(" | ".join(str(c) for c in row if c is not None))
            sample = "\n".join(lines)
        else:
            sample = content.decode("utf-8", "ignore")[:3000]
        return _detect_call([{"type": "text", "text": head + sample + "\n\n" + DETECTION_PROMPT}])
    except Exception:
        return empty


# Scripts that signal a non-English product name worth translating: CJK (Chinese /
# Japanese kanji), kana, Hangul, Cyrillic, Thai. HK vet catalogues are mostly Chinese.
_NON_LATIN = re.compile(r'[㐀-鿿぀-ヿ가-힯Ѐ-ӿ฀-๿]')


def _needs_translation(s: str) -> bool:
    return bool(s) and bool(_NON_LATIN.search(s))


TRANSLATION_PROMPT = """Translate each veterinary / pet-product name below into natural English, the way a Hong Kong pet shop would list it on its English store.
KEEP brand names, product/model codes, dosages, sizes, volumes and numbers EXACTLY (e.g. "10ml", "16mg", "x100", "2kg"). Translate only the descriptive words. If a line is already English, return it unchanged.
Return ONLY a JSON array: [{"i": <index>, "en": "<english name>"}], one object per input line using the SAME index number. No prose.

Names:
%s"""


def translate_to_english(items: list[dict]) -> list[dict]:
    """Translate non-English product descriptions to English in place, preserving the source
    in `original_description`. Only descriptions containing non-Latin script are sent to the
    model, so English catalogues incur no cost. Degrades to a no-op on any failure."""
    if not ANTHROPIC_API_KEY or not items:
        return items
    targets = [it for it in items
               if isinstance(it, dict) and _needs_translation(it.get("description") or "")]
    if not targets:
        return items
    import anthropic
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
    BATCH = 40
    for base in range(0, len(targets), BATCH):
        batch = targets[base:base + BATCH]
        lines = "\n".join(f'{n}. {it.get("description")}' for n, it in enumerate(batch))
        try:
            msg = client.messages.create(
                model="claude-haiku-4-5-20251001", max_tokens=MAX_TOKENS,
                messages=[{"role": "user", "content": TRANSLATION_PROMPT % lines}])
            out = _loads_json_array(msg.content[0].text)
        except Exception:
            continue
        by_i: dict[int, str] = {}
        for o in out:
            try:
                by_i[int(o.get("i"))] = (o.get("en") or "").strip()
            except (TypeError, ValueError):
                continue
        for n, it in enumerate(batch):
            en = by_i.get(n)
            if en and en != it.get("description"):
                it["original_description"] = it.get("description")
                it["description"] = en
    return items


def extract(content: bytes, filename: str, content_type: str, contract=None) -> tuple[list[dict], str]:
    """
    Main entry point. Dispatches to the right extractor based on file type, then translates
    any non-English descriptions to English (keeping the original). Returns (items, format).
    `contract` is an optional legacy extraction mapping. None means today's
    generic extraction path is used unchanged.
    """
    name_lower = filename.lower()
    ct_lower   = content_type.lower()

    if name_lower.endswith(('.xlsx', '.xls')) or 'spreadsheet' in ct_lower or 'excel' in ct_lower:
        items, fmt = extract_from_excel(content, filename), 'xlsx'
    elif name_lower.endswith('.csv') or 'csv' in ct_lower:
        items, fmt = _stub_result(filename), 'csv'
        if ANTHROPIC_API_KEY:
            try:
                got = _call_claude_text(content.decode('utf-8', errors='replace'), contract)
                items = got if isinstance(got, list) else []
            except Exception:
                pass
    elif name_lower.endswith('.pdf') or 'pdf' in ct_lower:
        items, fmt = extract_from_pdf(content, filename, contract), 'pdf'
    elif name_lower.endswith(('.jpg', '.jpeg')) or 'jpeg' in ct_lower:
        items, fmt = extract_from_image(content, 'image/jpeg', filename, contract), 'jpeg'
    elif name_lower.endswith('.png') or 'png' in ct_lower:
        items, fmt = extract_from_image(content, 'image/png', filename, contract), 'jpeg'
    else:
        items, fmt = _stub_result(filename), 'unknown'
        if ANTHROPIC_API_KEY:
            try:
                got = _call_claude_text(content.decode('utf-8', errors='replace'), contract)
                items = got if isinstance(got, list) else []
            except Exception:
                pass

    return translate_to_english(items), fmt
