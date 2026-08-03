"""Typed, source-located extraction for Staging after Raw has completed.

This module deliberately does not parse catalogue evidence into product fields.
It records what was observed and where it was observed. Semantic interpretation
belongs in Intermediate after Staging evidence persistence.
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import os
import platform
import re
import time
from concurrent.futures import ThreadPoolExecutor
from decimal import Decimal
from enum import Enum
from pathlib import Path
from typing import Any, Literal

import openpyxl
import pypdf
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, ConfigDict, Field, field_validator, model_validator

from schemas.catalogue_pipeline.enums import ExtractionMethod, SourceFormat
from schemas.catalogue_pipeline.extracted_evidence_v1 import BoundingBox, RawCell, SourceLocation
from services.catalogue_vision_provider import (
    VisionExtractionFailure,
    VisionResponse,
    active_provider,
)

# The vision provider owns the model id, the request shape and the retry
# classification; this module owns the prompt and what is done with the answer.
# Names kept for the pipeline's internal vocabulary.
_VisionResponse = VisionResponse
_VisionExtractionFailure = VisionExtractionFailure


# A multi-page PDF is N sequential provider round-trips otherwise; run the
# per-page vision calls concurrently, bounded here (env-overridable).
_VISION_CONCURRENCY = max(1, int(os.environ.get("CATALOGUE_VISION_CONCURRENCY", "6")))
# Retry a failed page inside the same extraction attempt so successful pages
# are not re-sent merely because one provider call was throttled.
_VISION_UNIT_RETRIES = max(0, int(os.environ.get("CATALOGUE_VISION_UNIT_RETRIES", "2")))
# Linear per-attempt backoff before a unit retry (seconds x attempt number).
_VISION_RETRY_BACKOFF_SECONDS = max(0.0, float(os.environ.get("CATALOGUE_VISION_RETRY_BACKOFF_SECONDS", "20")))
# Sparse-page completeness heuristic: warn when an evidence-bearing page yields
# <= MAX_ROWS while a sibling page in the same document yields >= SIBLING_MIN.
_SPARSE_PAGE_MAX_ROWS = max(0, int(os.environ.get("CATALOGUE_VISION_SPARSE_PAGE_MAX_ROWS", "2")))
_SPARSE_PAGE_SIBLING_MIN = max(1, int(os.environ.get("CATALOGUE_VISION_SPARSE_SIBLING_MIN", "8")))


class ExtractionStatus(str, Enum):
    """Completeness of one source extraction attempt."""

    COMPLETE = "COMPLETE"
    PARTIAL = "PARTIAL"
    FAILED = "FAILED"


class ExtractionError(BaseModel):
    """Sanitized operational error attached to an extraction result."""

    model_config = ConfigDict(extra="forbid", frozen=True)

    code: str = Field(..., min_length=1)
    message: str = Field(..., min_length=1)
    unit_key: str | None = None
    provider: str | None = None
    retryable: bool = False


class ExtractionUnitStatus(str, Enum):
    """Terminal outcome for one independently extractable source unit."""

    EVIDENCE_CAPTURED = "EVIDENCE_CAPTURED"
    NO_CATALOGUE_EVIDENCE = "NO_CATALOGUE_EVIDENCE"
    FAILED_RETRYABLE = "FAILED_RETRYABLE"
    FAILED_PERMANENT = "FAILED_PERMANENT"


class ExtractionUnitOutcome(BaseModel):
    """Completeness accounting for one page, worksheet, document or image."""

    model_config = ConfigDict(extra="forbid", frozen=True)

    unit_key: str = Field(..., min_length=1)
    status: ExtractionUnitStatus
    observation_count: int = Field(0, ge=0)
    attempt_count: int = Field(1, ge=1)
    provider: str | None = None
    provider_request_id: str | None = None
    error_code: str | None = None
    message: str | None = None

    @model_validator(mode="after")
    def _status_matches_count(self):
        if self.status == ExtractionUnitStatus.EVIDENCE_CAPTURED and self.observation_count <= 0:
            raise ValueError("EVIDENCE_CAPTURED requires at least one observation")
        if self.status != ExtractionUnitStatus.EVIDENCE_CAPTURED and self.observation_count:
            raise ValueError("Only EVIDENCE_CAPTURED may carry observations")
        if self.status in {
            ExtractionUnitStatus.FAILED_RETRYABLE,
            ExtractionUnitStatus.FAILED_PERMANENT,
        } and not self.error_code:
            raise ValueError("Failed extraction units require an error code")
        return self


class ExtractedEvidence(BaseModel):
    """One verbatim, source-located observation ready for Raw persistence."""

    model_config = ConfigDict(extra="forbid", frozen=True, protected_namespaces=())

    observation_key: str = Field(..., min_length=1)
    source_location: SourceLocation
    raw_text: str | None = None
    raw_cells: tuple[RawCell, ...] = ()
    extraction_method: ExtractionMethod
    provider: str | None = None
    provider_version: str | None = None
    provider_request_id: str | None = None
    model: str | None = None
    model_version: str | None = None
    confidence: Decimal | None = Field(None, ge=Decimal("0"), le=Decimal("1"))
    warnings: tuple[str, ...] = ()
    source_metadata: dict[str, Any] = Field(default_factory=dict)

    @field_validator("confidence", mode="before")
    @classmethod
    def _confidence_not_float(cls, value):
        if isinstance(value, float):
            raise ValueError("confidence must be a decimal string, integer, or Decimal")
        return value

    @model_validator(mode="after")
    def _requires_verbatim_evidence(self):
        has_text = bool(self.raw_text and self.raw_text.strip())
        has_cells = any(cell.raw_value is not None and str(cell.raw_value).strip() for cell in self.raw_cells)
        if not has_text and not has_cells:
            raise ValueError("ExtractedEvidence requires raw_text or at least one non-empty raw cell")
        return self


class ExtractionResult(BaseModel):
    """Typed extraction envelope with explicit completeness accounting.

    A source unit (PDF page, XLSX worksheet, CSV parse, image) only counts as
    completed when it was fully observed or the provider EXPLICITLY classified
    it as containing no catalogue evidence (``empty_units``). "The provider
    call returned" is never equated with "the unit was completely observed".
    """

    model_config = ConfigDict(extra="forbid", frozen=True)

    status: ExtractionStatus
    source_format: SourceFormat
    observations: tuple[ExtractedEvidence, ...] = ()
    units_attempted: int = Field(..., ge=0)
    units_completed: int = Field(..., ge=0)
    empty_units: int = Field(0, ge=0, description="Units explicitly classified as containing no catalogue evidence.")
    unit_outcomes: tuple[ExtractionUnitOutcome, ...] = ()
    warnings: tuple[str, ...] = ()
    errors: tuple[ExtractionError, ...] = ()

    @model_validator(mode="after")
    def _status_matches_contents(self):
        if self.units_completed > self.units_attempted:
            raise ValueError("units_completed cannot exceed units_attempted")
        if self.empty_units > self.units_completed:
            raise ValueError("empty_units cannot exceed units_completed")
        if self.unit_outcomes:
            if len(self.unit_outcomes) != self.units_attempted:
                raise ValueError("unit_outcomes must account for every attempted unit")
            keys = [item.unit_key for item in self.unit_outcomes]
            if len(keys) != len(set(keys)):
                raise ValueError("unit_outcome keys must be unique")
            completed = sum(
                item.status in {
                    ExtractionUnitStatus.EVIDENCE_CAPTURED,
                    ExtractionUnitStatus.NO_CATALOGUE_EVIDENCE,
                }
                for item in self.unit_outcomes
            )
            empty = sum(
                item.status == ExtractionUnitStatus.NO_CATALOGUE_EVIDENCE
                for item in self.unit_outcomes
            )
            if completed != self.units_completed or empty != self.empty_units:
                raise ValueError("unit_outcomes do not match completeness counters")
        if self.status == ExtractionStatus.COMPLETE:
            if not self.observations and not self.empty_units:
                raise ValueError("COMPLETE extraction requires observations or explicit empty-unit accounting")
            if self.errors or self.units_completed != self.units_attempted:
                raise ValueError("COMPLETE extraction cannot contain errors or incomplete units")
        elif self.status == ExtractionStatus.PARTIAL:
            if not (self.observations or self.empty_units) or not self.errors:
                raise ValueError("PARTIAL extraction requires observed/empty units and errors")
        elif self.status == ExtractionStatus.FAILED:
            if self.observations or not self.errors:
                raise ValueError("FAILED extraction requires errors and cannot contain observations")
        return self


class _VisionRow(BaseModel):
    """One page row in the COMPACT provider envelope.

    Tabular rows carry ``cells`` — verbatim values aligned by position to the
    page-level ``columns`` array (null for an empty cell; a short tail is
    padded as empty). Non-tabular lines carry ``text``. Emitting the column
    headings once per page instead of once per cell cuts provider output
    tokens roughly in half on dense bilingual tables — the dominant cost.
    """

    model_config = ConfigDict(extra="forbid")

    cells: tuple[str | None, ...] = ()
    text: str | None = None
    box: tuple[Decimal, Decimal, Decimal, Decimal] | None = None
    confidence: Decimal | None = Field(None, ge=Decimal("0"), le=Decimal("1"))

    @field_validator("cells", mode="before")
    @classmethod
    def _cells_are_verbatim_strings(cls, value):
        for item in value or ():
            if item is not None and not isinstance(item, str):
                raise ValueError("vision cell values must be verbatim strings")
        return value

    @model_validator(mode="after")
    def _requires_evidence(self):
        has_text = bool(self.text and self.text.strip())
        has_cells = any(cell is not None and cell.strip() for cell in self.cells)
        if not has_text and not has_cells:
            raise ValueError("vision row requires verbatim text or cells")
        return self


class _VisionTable(BaseModel):
    """One visual table and its own printed header family."""

    model_config = ConfigDict(extra="forbid")

    columns: tuple[str, ...]
    rows: tuple[_VisionRow, ...]
    # The banner printed ACROSS the table, above its column headings. For some
    # suppliers it is the only place a product's name appears: Alfamedic's
    # suture pages give each row a code, gauge, needle and price, and name the
    # material once, in the band over the block. Recorded verbatim against the
    # rows it spans, because a banner in text_observations is unattached to
    # anything and cannot be recovered later.
    section: str | None = None

    @model_validator(mode="after")
    def _rows_match_columns(self):
        # Columns with no rows is a heading whose rows are on the next page —
        # common once a table is emitted per section banner. It carries no
        # evidence, so the envelope drops it; refusing the whole page over it
        # would lose every row the page DOES have.
        if not self.columns:
            raise ValueError("vision table requires columns")
        for row in self.rows:
            if row.text is not None:
                raise ValueError("table rows must use cells; document text belongs in text_observations")
            if len(row.cells) > len(self.columns):
                raise ValueError("a row cannot carry more cells than its table columns")
        return self


class _VisionEnvelope(BaseModel):
    """Typed page-level provider outcome (compact form).

    ``page_outcome`` is REQUIRED: an empty row array is only acceptable when
    the provider explicitly classifies the page as containing no catalogue
    evidence. Empty-without-classification and evidence-with-empty-array are
    both malformed — they may hide truncation.
    """

    model_config = ConfigDict(extra="forbid")

    page_outcome: Literal["evidence", "no_catalogue_evidence"]
    tables: tuple[_VisionTable, ...] = ()
    text_observations: tuple[_VisionRow, ...] = ()
    # Backward-compatible reader for recorded compact-v1 envelopes. New provider
    # requests use tables/text_observations so one page can carry multiple
    # independent header families.
    columns: tuple[str, ...] = ()
    rows: tuple[_VisionRow, ...] = ()

    @model_validator(mode="after")
    def _outcome_matches_rows(self):
        rows_present = bool(self.rows or any(table.rows for table in self.tables))
        has_evidence = bool(rows_present or self.text_observations)
        if self.page_outcome == "evidence" and not has_evidence:
            raise ValueError("page_outcome 'evidence' requires table or text evidence")
        # The rule exists to catch a page classified empty that actually held
        # product ROWS, which would hide truncation. Document-level text is not
        # a product row: a page can genuinely have no catalogue lines and still
        # print an effective date or a policy note, and refusing it discards
        # both the classification and the text.
        if self.page_outcome == "no_catalogue_evidence" and rows_present:
            raise ValueError("page_outcome 'no_catalogue_evidence' cannot carry catalogue rows")
        if self.tables and (self.columns or self.rows):
            raise ValueError("use tables or legacy columns/rows, not both")
        for row in self.rows:
            if len(row.cells) > len(self.columns):
                raise ValueError("a row cannot carry more cells than declared columns")
        for row in self.text_observations:
            if row.cells or not row.text:
                raise ValueError("text_observations must contain text without cells")
        return self


class _VisionResponse(BaseModel):
    model_config = ConfigDict(extra="forbid", frozen=True)

    text: str
    request_id: str | None = None


VISION_EVIDENCE_PROMPT = """Extract only verbatim catalogue evidence from this source.

Identify each visually distinct catalogue row or product line. Do not interpret,
translate, normalize, calculate, split variants, infer product fields, or remove
duplicates. Preserve repeated rows at their separate source locations.

Return one JSON object with exactly this shape:
{
  "page_outcome": "evidence",
  "tables": [
    {
      "section": "banner printed across the table, verbatim; omit when absent",
      "columns": ["this table's printed headings, in printed order"],
      "rows": [
        {
          "cells": ["cell value exactly as printed, aligned to this table's columns, null when empty"],
          "confidence": "0.95"
        }
      ]
    }
  ],
  "text_observations": [
    {
      "text": "document-level evidence exactly as printed",
      "confidence": "0.9"
    }
  ]
}

page_outcome is REQUIRED and must be exactly one of:
- "evidence" — the page contains catalogue rows; rows must list every visible
  row without omission.
- "no_catalogue_evidence" — the page genuinely contains no catalogue rows
  and no relevant document-level text (blank page, pure artwork); tables and
  text_observations must be [].

Rules:
- Emit one tables entry per visually distinct table. Each table owns its
  printed headings; never force two different header families into one array.
- When a heading band spans the full width above a group of rows, emit that
  group as its own table and copy the band into that table's "section",
  verbatim. Some catalogues name a product only there — the rows below carry
  codes and specifications but no name.
- Product/offer rows belong in tables. Preserve commercially relevant
  document-level text — effective dates, price-basis policies, promotion
  periods, minimum-order rules and similar terms — in text_observations.
- Decorative titles, page numbers and contact lines may be omitted.
- Each tabular row's cells align by position to its table's columns; use null
  for an empty cell. A row never has more cells than its table's columns.
- Copy every value exactly as printed, including bilingual text, symbols and
  units. Never output computed, translated or normalized values.
- A row may optionally include "box": [x, y, width, height] in pixels.
- Use strings for confidence values. Return only the JSON object, without
  Markdown fences.
"""


def extract_evidence(content: bytes, filename: str, content_type: str) -> ExtractionResult:
    """Extract verbatim observations without performing semantic product parsing."""

    source_kind, source_format = _classify_source(filename, content_type)
    if not content:
        return _failed_result(
            source_format,
            code="EMPTY_SOURCE",
            message="Source content is empty",
        )
    if source_kind == "xlsx":
        return _extract_spreadsheet(content)
    if source_kind == "xls":
        return _failed_result(
            source_format,
            code="UNSUPPORTED_LEGACY_XLS",
            message="Legacy .xls files are not supported by the configured spreadsheet reader",
        )
    if source_kind == "csv":
        return _extract_csv(content)
    if source_kind == "pdf":
        return _extract_pdf(content)
    if source_kind in {"jpeg", "png"}:
        media_type = "image/png" if source_kind == "png" else "image/jpeg"
        return _extract_image(content, media_type=media_type, source_format=source_format)
    if source_kind == "text":
        return _extract_text(content, source_format=source_format)
    return _failed_result(
        source_format,
        code="UNSUPPORTED_SOURCE_FORMAT",
        message=f"Unsupported catalogue source type for {Path(filename).name or 'upload'}",
    )


def _extract_spreadsheet(content: bytes) -> ExtractionResult:
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=False)
    except Exception:
        return _failed_result(
            SourceFormat.SPREADSHEET,
            code="MALFORMED_SPREADSHEET",
            message="Spreadsheet source could not be read",
        )
    # Second read with data_only=True: for formula cells the workbook's CACHED
    # displayed value (what the supplier saw) becomes raw_value, with the
    # formula preserved verbatim on the cell. A workbook never opened by a
    # spreadsheet app has no cache — raw_value then stays the formula string.
    cached_values: dict[tuple[str, str], Any] = {}
    try:
        cached_workbook = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        try:
            for cached_sheet in cached_workbook.worksheets:
                for cached_row in cached_sheet.iter_rows():
                    for cached_cell in cached_row:
                        if cached_cell.value is not None:
                            cached_values[(cached_sheet.title, cached_cell.coordinate)] = cached_cell.value
        finally:
            cached_workbook.close()
    except Exception:
        cached_values = {}

    observations: list[ExtractedEvidence] = []
    warnings: list[str] = []
    errors: list[ExtractionError] = []
    unit_outcomes: list[ExtractionUnitOutcome] = []
    completed = 0
    sheets = tuple(workbook.worksheets)
    try:
        for sheet in sheets:
            observed_rows = 0
            header: dict[int, str | None] = {}
            try:
                for row in sheet.iter_rows():
                    non_empty = [cell for cell in row if cell.value is not None and str(cell.value).strip()]
                    if not non_empty:
                        continue
                    observed_rows += 1
                    if not header:
                        header = {cell.column: (str(cell.value).strip() or None) for cell in non_empty}
                    first_column = min(cell.column for cell in non_empty)
                    last_column = max(cell.column for cell in non_empty)
                    row_number = non_empty[0].row
                    cell_range = (
                        f"{get_column_letter(first_column)}{row_number}:"
                        f"{get_column_letter(last_column)}{row_number}"
                    )
                    key = f"sheet:{sheet.title}:row:{row_number}"
                    raw_cells = tuple(
                        RawCell(
                            cell_reference=cell.coordinate,
                            row_number=cell.row,
                            column_index=cell.column,
                            column_name=header.get(cell.column),
                            raw_value=(
                                cached_values.get((sheet.title, cell.coordinate), cell.value)
                                if _is_formula_value(cell.value)
                                else cell.value
                            ),
                            formula=cell.value if _is_formula_value(cell.value) else None,
                        )
                        for cell in row
                    )
                    observations.append(
                        ExtractedEvidence(
                            observation_key=key,
                            source_location=SourceLocation(
                                sheet_name=sheet.title,
                                row_number=row_number,
                                cell_range=cell_range,
                                source_object_key=key,
                            ),
                            raw_cells=raw_cells,
                            extraction_method=ExtractionMethod.SPREADSHEET_CELL,
                            provider="openpyxl",
                            provider_version=openpyxl.__version__,
                        )
                    )
            except Exception:
                error = ExtractionError(
                        code="SPREADSHEET_SHEET_READ_ERROR",
                        message="One spreadsheet sheet could not be read completely",
                        unit_key=f"sheet:{sheet.title}",
                        provider="openpyxl",
                    )
                errors.append(error)
                unit_outcomes.append(
                    ExtractionUnitOutcome(
                        unit_key=f"sheet:{sheet.title}",
                        status=ExtractionUnitStatus.FAILED_PERMANENT,
                        provider="openpyxl",
                        error_code=error.code,
                        message=error.message,
                    )
                )
                continue
            completed += 1
            if not observed_rows:
                warnings.append(f"sheet {sheet.title!r} contained no non-empty rows")
                unit_outcomes.append(
                    ExtractionUnitOutcome(
                        unit_key=f"sheet:{sheet.title}",
                        status=ExtractionUnitStatus.NO_CATALOGUE_EVIDENCE,
                        provider="openpyxl",
                    )
                )
            else:
                unit_outcomes.append(
                    ExtractionUnitOutcome(
                        unit_key=f"sheet:{sheet.title}",
                        status=ExtractionUnitStatus.EVIDENCE_CAPTURED,
                        observation_count=observed_rows,
                        provider="openpyxl",
                    )
                )
    finally:
        workbook.close()

    return _build_result(
        SourceFormat.SPREADSHEET,
        observations=observations,
        units_attempted=len(sheets),
        units_completed=completed,
        empty_units=sum(
            item.status == ExtractionUnitStatus.NO_CATALOGUE_EVIDENCE
            for item in unit_outcomes
        ),
        unit_outcomes=unit_outcomes,
        warnings=warnings,
        errors=errors,
    )


def _is_formula_value(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _header_labels(rows: list[list[str]]) -> dict[int, str | None]:
    """First non-empty row's values as column labels, keyed by 1-based column index.

    Structural inference only (no semantics): lets the supplier contract map
    each data cell by its column heading. The header row itself is recognised
    and skipped downstream (its cells repeat the contract's source columns).
    """

    for row in rows:
        if any(str(value).strip() for value in row):
            return {index: (str(value).strip() or None) for index, value in enumerate(row, start=1)}
    return {}


def _extract_csv(content: bytes) -> ExtractionResult:
    try:
        text, encoding = _decode_delimited_text(content)
    except UnicodeDecodeError:
        return _failed_result(
            SourceFormat.CSV,
            code="UNSUPPORTED_TEXT_ENCODING",
            message="CSV source is not valid UTF-8 or Big5 text",
        )
    try:
        dialect = _sniff_dialect(text)
        rows = list(csv.reader(io.StringIO(text), dialect))
    except csv.Error:
        return _failed_result(
            SourceFormat.CSV,
            code="MALFORMED_CSV",
            message="CSV source could not be parsed",
        )

    header = _header_labels(rows)
    observations: list[ExtractedEvidence] = []
    for row_number, row in enumerate(rows, start=1):
        if not any(value.strip() for value in row):
            continue
        key = f"csv:row:{row_number}"
        last_column = max(1, len(row))
        observations.append(
            ExtractedEvidence(
                observation_key=key,
                source_location=SourceLocation(
                    row_number=row_number,
                    cell_range=f"A{row_number}:{get_column_letter(last_column)}{row_number}",
                    source_object_key=key,
                ),
                raw_cells=tuple(
                    RawCell(
                        cell_reference=f"{get_column_letter(column_index)}{row_number}",
                        row_number=row_number,
                        column_index=column_index,
                        column_name=header.get(column_index),
                        raw_value=value,
                    )
                    for column_index, value in enumerate(row, start=1)
                ),
                extraction_method=ExtractionMethod.SPREADSHEET_CELL,
                provider="python-csv",
                provider_version=platform.python_version(),
                source_metadata={"encoding": encoding, "delimiter": dialect.delimiter},
            )
        )

    return _build_result(
        SourceFormat.CSV,
        observations=observations,
        units_attempted=1,
        units_completed=1,
        empty_units=0 if observations else 1,
        unit_outcomes=[
            ExtractionUnitOutcome(
                unit_key="csv:1",
                status=(
                    ExtractionUnitStatus.EVIDENCE_CAPTURED
                    if observations
                    else ExtractionUnitStatus.NO_CATALOGUE_EVIDENCE
                ),
                observation_count=len(observations),
                provider="python-csv",
            )
        ],
    )


def _extract_text(content: bytes, *, source_format: SourceFormat) -> ExtractionResult:
    try:
        text, encoding = _decode_delimited_text(content)
    except UnicodeDecodeError:
        return _failed_result(
            source_format,
            code="UNSUPPORTED_TEXT_ENCODING",
            message="Text source is not valid UTF-8 or Big5 text",
        )
    lines = text.splitlines()
    observations = [
        ExtractedEvidence(
            observation_key=f"text:line:{line_number}",
            source_location=SourceLocation(
                source_object_key=f"text:line:{line_number}",
            ),
            raw_text=line,
            extraction_method=ExtractionMethod.OTHER,
            provider="python-text",
            provider_version=platform.python_version(),
            source_metadata={"encoding": encoding, "line_number": line_number},
        )
        for line_number, line in enumerate(lines, start=1)
        if line.strip()
    ]
    return _build_result(
        source_format,
        observations=observations,
        units_attempted=1,
        units_completed=1,
        empty_units=0 if observations else 1,
        unit_outcomes=[
            ExtractionUnitOutcome(
                unit_key="text:1",
                status=(
                    ExtractionUnitStatus.EVIDENCE_CAPTURED
                    if observations
                    else ExtractionUnitStatus.NO_CATALOGUE_EVIDENCE
                ),
                observation_count=len(observations),
                provider="python-text",
            )
        ],
    )


def _extract_pdf(content: bytes) -> ExtractionResult:
    try:
        reader = pypdf.PdfReader(io.BytesIO(content))
        # Owner-locked PDFs (empty user password) already passed the raw stage;
        # decrypt explicitly rather than relying on pypdf's lazy empty-password
        # fallback during page access.
        if reader.is_encrypted and not reader.decrypt(""):
            return _failed_result(
                SourceFormat.PDF,
                code="MALFORMED_PDF",
                message="PDF source is password protected",
            )
    except Exception:
        return _failed_result(
            SourceFormat.PDF,
            code="MALFORMED_PDF",
            message="PDF source could not be read",
        )

    pages = list(reader.pages)
    observations: list[ExtractedEvidence] = []
    warnings: list[str] = []
    errors: list[ExtractionError] = []
    unit_outcomes: list[ExtractionUnitOutcome] = []
    completed = 0
    empty_units = 0

    if not pages:
        warnings.append("PDF contained no pages")
        return _build_result(
            SourceFormat.PDF,
            observations=observations,
            units_attempted=0,
            units_completed=0,
            unit_outcomes=[],
        )

    # Every catalogue page is extracted via the vision provider so tables yield
    # column-labeled cells the supplier contract can map deterministically. The
    # PDF text layer is not used as evidence (it linearizes multi-column tables
    # into per-cell lines the contract cannot map to fields).
    provider = active_provider()
    if not provider.is_configured():
        errors = [
            ExtractionError(
                code="EXTRACTION_CONFIGURATION_ERROR",
                message=f"PDF evidence extraction requires a configured {provider.name} vision provider",
                unit_key=f"page:{page_number}",
                provider=provider.name,
            )
            for page_number in range(1, len(pages) + 1)
        ]
        return _build_result(
            SourceFormat.PDF,
            observations=observations,
            units_attempted=len(pages),
            units_completed=0,
            unit_outcomes=[
                ExtractionUnitOutcome(
                    unit_key=error.unit_key or f"page:{index}",
                    status=ExtractionUnitStatus.FAILED_PERMANENT,
                    provider=error.provider,
                    error_code=error.code,
                    message=error.message,
                )
                for index, error in enumerate(errors, start=1)
            ],
            errors=errors,
        )

    # Prepare per-page bytes sequentially (pypdf is not thread-safe), then run
    # the vision round-trips concurrently — a multi-page PDF is otherwise N
    # sequential provider calls.
    page_payloads: list[tuple[int, bytes]] = []
    for page_number, page in enumerate(pages, start=1):
        try:
            page_payloads.append((page_number, _single_page_pdf_bytes(page)))
        except Exception:
            error = ExtractionError(
                code="SOURCE_PAGE_READ_ERROR",
                message="PDF page could not be prepared for vision extraction",
                unit_key=f"page:{page_number}",
                provider="pypdf",
            )
            errors.append(error)
            unit_outcomes.append(
                ExtractionUnitOutcome(
                    unit_key=f"page:{page_number}",
                    status=ExtractionUnitStatus.FAILED_PERMANENT,
                    provider="pypdf",
                    error_code=error.code,
                    message=error.message,
                )
            )

    def _extract_page(page_number: int, page_bytes: bytes):
        page_key = f"page:{page_number}"
        attempt_count = 0
        while True:
            attempt_count += 1
            try:
                response = _call_vision(page_bytes, media_type="application/pdf")
                page_observations, page_outcome = _vision_observations(
                    response,
                    extraction_method=ExtractionMethod.MODEL_VISION,
                    unit_key=page_key,
                    page_number=page_number,
                )
                return (
                    page_number,
                    page_observations,
                    page_outcome,
                    None,
                    attempt_count,
                )
            except _VisionExtractionFailure as exc:
                if exc.retryable and attempt_count <= _VISION_UNIT_RETRIES:
                    # Back off before re-calling: retryable failures are mostly
                    # per-minute throttles (429) — immediate retries land inside
                    # the same window and burn every attempt in seconds.
                    time.sleep(_VISION_RETRY_BACKOFF_SECONDS * attempt_count)
                    continue
                return page_number, [], None, ExtractionError(
                    code=exc.code,
                    message=exc.public_message,
                    unit_key=page_key,
                    provider=provider.name,
                    retryable=exc.retryable,
                ), attempt_count
            except Exception:
                return page_number, [], None, ExtractionError(
                    code="SOURCE_PAGE_READ_ERROR",
                    message="PDF page could not be prepared for vision extraction",
                    unit_key=page_key,
                    provider="pypdf",
                ), attempt_count

    workers = min(len(page_payloads), _VISION_CONCURRENCY) or 1
    with ThreadPoolExecutor(max_workers=workers) as pool:
        page_results = list(pool.map(lambda payload: _extract_page(*payload), page_payloads))

    # Assemble in page order so observation identity/ordering stays deterministic.
    for page_number, page_observations, page_outcome, error, attempt_count in sorted(
        page_results,
        key=lambda item: item[0],
    ):
        if error is not None:
            errors.append(error)
            unit_outcomes.append(
                ExtractionUnitOutcome(
                    unit_key=error.unit_key or f"page:{page_number}",
                    status=(
                        ExtractionUnitStatus.FAILED_RETRYABLE
                        if error.retryable
                        else ExtractionUnitStatus.FAILED_PERMANENT
                    ),
                    provider=error.provider,
                    attempt_count=attempt_count,
                    error_code=error.code,
                    message=error.message,
                )
            )
            continue
        completed += 1
        if page_outcome == "no_catalogue_evidence":
            warnings.append(f"page {page_number}: provider classified page as containing no catalogue evidence")
            empty_units += 1
            unit_outcomes.append(
                ExtractionUnitOutcome(
                    unit_key=f"page:{page_number}",
                    status=ExtractionUnitStatus.NO_CATALOGUE_EVIDENCE,
                    provider=provider.name,
                    attempt_count=attempt_count,
                )
            )
        else:
            unit_outcomes.append(
                ExtractionUnitOutcome(
                    unit_key=f"page:{page_number}",
                    status=ExtractionUnitStatus.EVIDENCE_CAPTURED,
                    observation_count=len(page_observations),
                    provider=provider.name,
                    attempt_count=attempt_count,
                    provider_request_id=(
                        page_observations[0].provider_request_id
                        if page_observations
                        else None
                    ),
                )
            )
        observations.extend(page_observations)

    # Completeness heuristic: a provider response can be valid, evidence-bearing
    # and still silently missing rows (the observed thinking-level failure
    # mode). A page yielding 1-2 rows in a document whose sibling pages yield
    # many is suspicious — surface a reviewable warning, never a failure.
    captured_counts = {
        outcome.unit_key: outcome.observation_count
        for outcome in unit_outcomes
        if outcome.status == ExtractionUnitStatus.EVIDENCE_CAPTURED
    }
    if captured_counts:
        densest = max(captured_counts.values())
        if densest >= _SPARSE_PAGE_SIBLING_MIN:
            for unit_key, count in sorted(captured_counts.items()):
                if count <= _SPARSE_PAGE_MAX_ROWS:
                    warnings.append(
                        f"{unit_key}: suspiciously sparse extraction ({count} row(s) while a sibling "
                        f"page yielded {densest}); verify page completeness before relying on this page"
                    )

    return _build_result(
        SourceFormat.PDF,
        observations=observations,
        units_attempted=len(pages),
        units_completed=completed,
        empty_units=empty_units,
        unit_outcomes=unit_outcomes,
        warnings=warnings,
        errors=errors,
    )


def _extract_image(
    content: bytes,
    *,
    media_type: str,
    source_format: SourceFormat,
) -> ExtractionResult:
    provider = active_provider()
    if not provider.is_configured():
        return _failed_result(
            source_format,
            code="EXTRACTION_CONFIGURATION_ERROR",
            message=f"Image evidence extraction requires a configured {provider.name} vision provider",
            unit_key="image:1",
            provider=provider.name,
        )
    try:
        response = _call_vision(content, media_type=media_type)
        observations, page_outcome = _vision_observations(
            response,
            extraction_method=ExtractionMethod.MODEL_VISION,
            unit_key="image:1",
            page_number=None,
        )
    except _VisionExtractionFailure as exc:
        return _failed_result(
            source_format,
            code=exc.code,
            message=exc.public_message,
            unit_key="image:1",
            provider=provider.name,
            retryable=exc.retryable,
            units_attempted=1,
        )
    warnings: list[str] = []
    empty_units = 0
    if page_outcome == "no_catalogue_evidence":
        warnings.append("image:1: provider classified image as containing no catalogue evidence")
        empty_units = 1
    return _build_result(
        source_format,
        observations=observations,
        units_attempted=1,
        units_completed=1,
        empty_units=empty_units,
        unit_outcomes=[
            ExtractionUnitOutcome(
                unit_key="image:1",
                status=(
                    ExtractionUnitStatus.NO_CATALOGUE_EVIDENCE
                    if page_outcome == "no_catalogue_evidence"
                    else ExtractionUnitStatus.EVIDENCE_CAPTURED
                ),
                observation_count=len(observations),
                provider=provider.name,
                provider_request_id=(
                    observations[0].provider_request_id if observations else None
                ),
            )
        ],
        warnings=warnings,
    )


def _pdf_text_observations(text: str, *, page_number: int) -> list[ExtractedEvidence]:
    observations: list[ExtractedEvidence] = []
    for line_number, line in enumerate(text.splitlines(), start=1):
        if not line.strip():
            continue
        key = f"page:{page_number}:line:{line_number}"
        observations.append(
            ExtractedEvidence(
                observation_key=key,
                source_location=SourceLocation(
                    page_number=page_number,
                    source_object_key=key,
                ),
                raw_text=line,
                extraction_method=ExtractionMethod.PDF_TEXT,
                provider="pypdf",
                provider_version=pypdf.__version__,
            )
        )
    return observations


def _vision_observations(
    response: _VisionResponse,
    *,
    extraction_method: ExtractionMethod,
    unit_key: str,
    page_number: int | None,
) -> tuple[list[ExtractedEvidence], Literal["evidence", "no_catalogue_evidence"]]:
    """Validate the typed provider envelope and mint stable observation keys.

    Identity policy: vision observation keys are derived from the observed
    CONTENT plus SOURCE LOCATION (sha256 of raw_text/raw_cells/bounding box),
    with a per-digest occurrence ordinal — so identical evidence keeps the
    same identity across provider retries regardless of response ORDER, and
    legitimate duplicate rows remain distinct (different bounding boxes yield
    different digests; byte-identical duplicates get interchangeable
    ordinals). Materially changed provider output yields different digests
    and is surfaced downstream as an idempotency conflict rather than
    silently corrupting persisted evidence.
    """

    provider = active_provider()
    try:
        payload = _strict_json_object(response.text)
        envelope = _VisionEnvelope.model_validate(payload)
    except (json.JSONDecodeError, ValueError) as exc:
        raise _VisionExtractionFailure(
            code="MALFORMED_PROVIDER_RESPONSE",
            public_message="Vision provider returned an invalid evidence envelope",
            # Model nondeterminism: a fresh generation usually produces a valid
            # envelope, so give the page its in-attempt unit retries.
            retryable=True,
        ) from exc

    if envelope.page_outcome == "no_catalogue_evidence":
        return [], "no_catalogue_evidence"

    try:
        observations: list[ExtractedEvidence] = []
        digest_counts: dict[str, int] = {}
        row_groups: list[tuple[tuple[str, ...], _VisionRow, str | None]] = []
        for table in envelope.tables:
            # An empty table is a banner whose rows sit on the next page.
            row_groups.extend((table.columns, row, table.section) for row in table.rows)
        # Recorded compact-v1 fixtures remain readable.
        row_groups.extend((envelope.columns, row, None) for row in envelope.rows)
        row_groups.extend(((), row, None) for row in envelope.text_observations)

        for columns, row, section in row_groups:
            digest = _observation_digest(columns, row)
            ordinal = digest_counts.get(digest, 0) + 1
            digest_counts[digest] = ordinal
            key = f"{unit_key}:obs:{digest}:{ordinal}"
            # Re-expand the compact row into the persisted evidence shape: the
            # page-level heading array becomes each cell's column_name, so
            # everything downstream (persistence, contract conformance,
            # replay idempotency) is untouched by the wire-format change.
            raw_cells = tuple(
                RawCell(
                    cell_reference=None,
                    row_number=None,
                    column_name=columns[index],
                    column_index=index + 1,
                    raw_value=value,
                )
                for index, value in enumerate(row.cells)
                if value is not None
            )
            bounding_box = (
                BoundingBox(x=row.box[0], y=row.box[1], width=row.box[2], height=row.box[3], unit="px")
                if row.box
                else None
            )
            observations.append(
                ExtractedEvidence(
                    observation_key=key,
                    source_location=SourceLocation(
                        page_number=page_number,
                        bounding_box=bounding_box,
                        source_object_key=key,
                    ),
                    raw_text=row.text,
                    raw_cells=raw_cells,
                    extraction_method=extraction_method,
                    provider=provider.name,
                    provider_request_id=response.request_id,
                    model=provider.model,
                    model_version=provider.model,
                    confidence=row.confidence,
                    source_metadata={"section": section} if section else {},
                )
            )
    except ValueError as exc:
        raise _VisionExtractionFailure(
            code="MALFORMED_PROVIDER_RESPONSE",
            public_message="Vision provider returned invalid source evidence",
            retryable=True,
        ) from exc
    return observations, "evidence"


def _observation_digest(columns: tuple[str, ...], row: _VisionRow) -> str:
    material = {
        "columns": list(columns[: len(row.cells)]),
        "cells": [None if value is None else str(value) for value in row.cells],
        "text": row.text,
        "box": [str(value) for value in row.box] if row.box else None,
    }
    return hashlib.sha256(json.dumps(material, sort_keys=True, ensure_ascii=False).encode()).hexdigest()[:16]


def _call_vision(content: bytes, *, media_type: str) -> _VisionResponse:
    """Ask the configured vision provider to read one page.

    The single point where catalogue evidence meets a model. Which provider
    answers is an environment decision; everything downstream — the envelope,
    the observation identity, the retry policy — is identical either way.
    """
    return active_provider().call(content, media_type=media_type, prompt=VISION_EVIDENCE_PROMPT)




def _single_page_pdf_bytes(page) -> bytes:
    writer = pypdf.PdfWriter()
    writer.add_page(page)
    output = io.BytesIO()
    writer.write(output)
    return output.getvalue()


def _strict_json_object(raw: str) -> dict[str, Any]:
    stripped = raw.strip()
    stripped = re.sub(r"^```(?:json)?\s*", "", stripped)
    stripped = re.sub(r"\s*```$", "", stripped)
    # Live JSON-mode responses occasionally append a stray closing brace or
    # bracket after an otherwise complete envelope (observed on real supplier
    # pages). Parse the FIRST JSON object and tolerate ONLY structural debris
    # (whitespace/brackets/braces) after it — any real trailing content still
    # fails, and the envelope's semantic validation runs unchanged after this.
    try:
        value = json.loads(stripped, parse_float=Decimal)
    except json.JSONDecodeError:
        value, end = json.JSONDecoder(parse_float=Decimal).raw_decode(stripped)
        remainder = stripped[end:].strip()
        if remainder and not re.fullmatch(r"[\s\]\}]*", remainder):
            raise
    if not isinstance(value, dict):
        raise ValueError("provider response must be one JSON object")
    return value


def _build_result(
    source_format: SourceFormat,
    *,
    observations: list[ExtractedEvidence],
    units_attempted: int,
    units_completed: int,
    empty_units: int = 0,
    unit_outcomes: list[ExtractionUnitOutcome] | None = None,
    warnings: list[str] | None = None,
    errors: list[ExtractionError] | None = None,
) -> ExtractionResult:
    warnings = list(warnings or [])
    errors = list(errors or [])
    # No observations, no explicit empty-unit classification and no recorded
    # errors can never read as success — that combination is NO_EVIDENCE.
    if not observations and not empty_units and not errors:
        errors.append(
            ExtractionError(
                code="NO_EVIDENCE",
                message="Extraction completed without any non-empty source observations",
            )
        )
    observed_or_empty = bool(observations) or empty_units > 0
    if observed_or_empty and errors:
        status = ExtractionStatus.PARTIAL
    elif observed_or_empty:
        status = ExtractionStatus.COMPLETE
    else:
        status = ExtractionStatus.FAILED
    return ExtractionResult(
        status=status,
        source_format=source_format,
        observations=tuple(observations),
        units_attempted=units_attempted,
        units_completed=units_completed,
        empty_units=empty_units,
        unit_outcomes=tuple(unit_outcomes or ()),
        warnings=tuple(warnings),
        errors=tuple(errors),
    )


def _failed_result(
    source_format: SourceFormat,
    *,
    code: str,
    message: str,
    unit_key: str | None = None,
    provider: str | None = None,
    retryable: bool = False,
    units_attempted: int = 0,
) -> ExtractionResult:
    outcomes = ()
    if units_attempted:
        outcomes = tuple(
            ExtractionUnitOutcome(
                unit_key=unit_key or f"unit:{index}",
                status=(
                    ExtractionUnitStatus.FAILED_RETRYABLE
                    if retryable
                    else ExtractionUnitStatus.FAILED_PERMANENT
                ),
                provider=provider,
                error_code=code,
                message=message,
            )
            for index in range(1, units_attempted + 1)
        )
    return ExtractionResult(
        status=ExtractionStatus.FAILED,
        source_format=source_format,
        observations=(),
        units_attempted=units_attempted,
        units_completed=0,
        unit_outcomes=outcomes,
        errors=(
            ExtractionError(
                code=code,
                message=message,
                unit_key=unit_key,
                provider=provider,
                retryable=retryable,
            ),
        ),
    )


def _decode_delimited_text(content: bytes) -> tuple[str, str]:
    for encoding in ("utf-8-sig", "big5"):
        try:
            return content.decode(encoding), encoding
        except UnicodeDecodeError:
            continue
    raise UnicodeDecodeError("catalogue", content, 0, len(content), "unsupported text encoding")


def _sniff_dialect(text: str):
    sample = text[:8192]
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        return csv.excel


def _classify_source(filename: str, content_type: str) -> tuple[str, SourceFormat]:
    suffix = Path(filename or "").suffix.lower()
    normalized_type = (content_type or "").split(";", 1)[0].strip().lower()
    if suffix == ".xls":
        return "xls", SourceFormat.SPREADSHEET
    if suffix == ".xlsx":
        return "xlsx", SourceFormat.SPREADSHEET
    if suffix in {".csv", ".tsv"}:
        return "csv", SourceFormat.CSV
    if suffix == ".pdf":
        return "pdf", SourceFormat.PDF
    if suffix in {".jpg", ".jpeg"}:
        return "jpeg", SourceFormat.IMAGE
    if suffix == ".png":
        return "png", SourceFormat.IMAGE
    if suffix == ".txt":
        return "text", SourceFormat.TEXT
    if normalized_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        return "xlsx", SourceFormat.SPREADSHEET
    if normalized_type == "application/vnd.ms-excel":
        return "xls", SourceFormat.SPREADSHEET
    if normalized_type in {"text/csv", "application/csv"}:
        return "csv", SourceFormat.CSV
    if normalized_type == "application/pdf":
        return "pdf", SourceFormat.PDF
    if normalized_type == "image/jpeg":
        return "jpeg", SourceFormat.IMAGE
    if normalized_type == "image/png":
        return "png", SourceFormat.IMAGE
    if normalized_type.startswith("text/"):
        return "text", SourceFormat.TEXT
    return "unknown", SourceFormat.OTHER


def _pdf_text_is_reliable(text: str) -> bool:
    chars = [character for character in text if not character.isspace()]
    if not chars:
        return False

    def expected(character: str) -> bool:
        codepoint = ord(character)
        return (
            0x20 <= codepoint <= 0x7E
            or 0x00A0 <= codepoint <= 0x00FF
            or 0x2000 <= codepoint <= 0x206F
            or 0x3000 <= codepoint <= 0x303F
            or 0x3400 <= codepoint <= 0x4DBF
            or 0x4E00 <= codepoint <= 0x9FFF
            or 0xFF00 <= codepoint <= 0xFFEF
        )

    suspicious = sum(1 for character in chars if not expected(character))
    return Decimal(suspicious) / Decimal(len(chars)) < Decimal("0.12")


# ── PDF page extraction policy (document-level, never semantic) ──────────────
# A page is classified by text QUALITY plus a STRUCTURAL image summary.
# Text-line count alone can never suppress vision: if material (or
# unknown-coverage) visual content may hold uncaptured evidence, vision runs.
#
#   no text                                   -> VISION      (NO_TEXT)
#   garbled text                              -> VISION      (UNRELIABLE_TEXT)
#   reliable text, no/decorative-only images  -> TEXT        (RELIABLE_TEXT_NO_MATERIAL_IMAGES)
#   reliable text + material image(s)         -> HYBRID      (TEXT_WITH_MATERIAL_IMAGES or
#                                                             SPARSE_TEXT_WITH_MATERIAL_IMAGES)
#   reliable text + undeterminable coverage   -> UNCERTAIN   (IMAGE_COVERAGE_UNKNOWN, vision runs)
#
# Structural metrics only — pixel dimensions of image XObjects, never decoded
# content, never business-field detection. Rendered coverage requires the
# content-stream CTM, which is deliberately not parsed: mid-size or
# dimensionless images are UNCERTAIN and vision runs rather than risking
# silent evidence loss. Vision does not guarantee absolute completeness; the
# decision, reason code and metrics are recorded honestly. Hybrid pages keep
# BOTH evidence sets (distinguishable by method/key/location, no dedup here);
# overlap resolution is interpretation's job downstream.

_MIN_MEANINGFUL_TEXT_LINES = 3
_DECORATIVE_MAX_DIMENSION_PX = 200      # small logos / icons
_MATERIAL_MIN_PIXEL_AREA = 250_000      # ~500x500+: scans, photographed tables

_NOISE_LINE_PATTERNS = (
    re.compile(r"^\d{1,4}$"),                                   # bare page number
    re.compile(r"^page\s*\d+(\s*(of|/)\s*\d+)?$", re.IGNORECASE),
    re.compile(r"^[-–—•·.*_=~|]+$"),                            # ruler / separator noise
    re.compile(r"^\d{1,2}[/.-]\d{1,2}[/.-]\d{2,4}$"),           # bare date footer
)


class PdfPageMode(str, Enum):
    TEXT = "text"
    VISION = "vision"
    HYBRID = "hybrid"
    UNCERTAIN = "uncertain"


class _PageImageSummary:
    __slots__ = ("image_count", "decorative_count", "material_count", "unknown_count", "largest_pixel_area", "unreadable")

    def __init__(self):
        self.image_count = 0
        self.decorative_count = 0
        self.material_count = 0
        self.unknown_count = 0
        self.largest_pixel_area = 0
        self.unreadable = False


class _PageDecision:
    __slots__ = ("mode", "keep_text", "vision_required", "reason", "metrics")

    def __init__(self, mode: PdfPageMode, keep_text: bool, vision_required: bool, reason: str, metrics: dict | None = None):
        self.mode = mode
        self.keep_text = keep_text
        self.vision_required = vision_required
        self.reason = reason
        self.metrics = metrics or {}


def _classify_pdf_page(page, page_text: str) -> _PageDecision:
    has_text = bool(page_text.strip())
    if not has_text:
        return _PageDecision(PdfPageMode.VISION, keep_text=False, vision_required=True, reason="NO_TEXT")
    if not _pdf_text_is_reliable(page_text):
        return _PageDecision(PdfPageMode.VISION, keep_text=False, vision_required=True, reason="UNRELIABLE_TEXT")

    images = _summarize_page_images(page)
    meaningful = _meaningful_line_count(page_text)
    metrics = {
        "meaningful_lines": meaningful,
        "image_count": images.image_count,
        "decorative_images": images.decorative_count,
        "material_images": images.material_count,
        "unknown_images": images.unknown_count,
        "largest_pixel_area": images.largest_pixel_area,
    }
    if images.unreadable or images.unknown_count:
        return _PageDecision(
            PdfPageMode.UNCERTAIN, keep_text=True, vision_required=True,
            reason="IMAGE_COVERAGE_UNKNOWN", metrics=metrics,
        )
    if images.material_count:
        reason = (
            "TEXT_WITH_MATERIAL_IMAGES"
            if meaningful >= _MIN_MEANINGFUL_TEXT_LINES
            else "SPARSE_TEXT_WITH_MATERIAL_IMAGES"
        )
        return _PageDecision(PdfPageMode.HYBRID, keep_text=True, vision_required=True, reason=reason, metrics=metrics)
    return _PageDecision(
        PdfPageMode.TEXT, keep_text=True, vision_required=False,
        reason="RELIABLE_TEXT_NO_MATERIAL_IMAGES", metrics=metrics,
    )


def _summarize_page_images(page) -> _PageImageSummary:
    """Structural image summary — XObject pixel dimensions only, no decoding.

    An image is decorative when both dimensions are small (logo/icon scale),
    material when its pixel area is scan/photo scale, and unknown when its
    dimensions are missing or in between (rendered coverage cannot be
    determined without parsing the content-stream CTM).
    """

    summary = _PageImageSummary()
    try:
        resources = page.get("/Resources")
        if resources is None:
            return summary
        xobjects = resources.get_object().get("/XObject")
        if xobjects is None:
            return summary
        xobjects = xobjects.get_object()
        for name in xobjects:
            candidate = xobjects[name].get_object()
            if candidate.get("/Subtype") != "/Image":
                continue
            summary.image_count += 1
            try:
                width = int(candidate["/Width"])
                height = int(candidate["/Height"])
            except Exception:
                summary.unknown_count += 1
                continue
            area = width * height
            summary.largest_pixel_area = max(summary.largest_pixel_area, area)
            if width <= _DECORATIVE_MAX_DIMENSION_PX and height <= _DECORATIVE_MAX_DIMENSION_PX:
                summary.decorative_count += 1
            elif area >= _MATERIAL_MIN_PIXEL_AREA:
                summary.material_count += 1
            else:
                summary.unknown_count += 1
    except Exception:
        summary.unreadable = True
    return summary


def _decision_warning(page_number: int, decision: _PageDecision) -> str | None:
    if decision.mode == PdfPageMode.HYBRID:
        return f"page {page_number}: {decision.reason} — using text and vision (hybrid)"
    if decision.mode == PdfPageMode.UNCERTAIN:
        return f"page {page_number}: {decision.reason} — image coverage undeterminable; using text and vision"
    if decision.reason == "UNRELIABLE_TEXT":
        return f"page {page_number}: text layer unreliable; using vision"
    return None


def _meaningful_line_count(text: str) -> int:
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    return sum(1 for line in lines if not _is_noise_line(line))


def _is_noise_line(line: str) -> bool:
    if len(line) < 4:
        return True
    return any(pattern.match(line) for pattern in _NOISE_LINE_PATTERNS)




__all__ = [
    "ExtractedEvidence",
    "ExtractionError",
    "ExtractionResult",
    "ExtractionStatus",
    "ExtractionUnitOutcome",
    "ExtractionUnitStatus",
    "extract_evidence",
]
